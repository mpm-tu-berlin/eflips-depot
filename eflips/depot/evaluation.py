# -*- coding: utf-8 -*-
"""Evaluation, plots and exports for after the simulation.

Requires Matplotlib version 3+

Plot by directly calling a method of DepotEvaluation after running the
simulation with data logging. Example call in the console or a script:
    simulation_host.depot_hosts[0].evaluation.nvehicles_total()

If a plot opens in PyCharm's SciView and is not displayed properly, then go to
Settings -> Tools -> Python Scientific -> uncheck "Show plots in toolwindow",
rerun the simulation, execute:
">>> import matplotlib"
">>> matplotlib.use('Qt5Agg')"
and plot again. Plots are now displayed in a separate window.
Source:
https://intellij-support.jetbrains.com/hc/en-us/community/posts/115000736584-
SciView-in-PyCharm-2017-3-reduces-functionality-of-Matplotlib

"""
import itertools
import locale
import operator
import os
import pprint as pp
import traceback
from collections import OrderedDict, Counter
from datetime import datetime, timedelta
from functools import reduce

import matplotlib.patches as mpatch
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
import pandas as pd
import xlsxwriter
from eflips.helperFunctions import cm2in
from eflips.settings import globalConstants
from matplotlib.font_manager import FontProperties
from openpyxl import Workbook
from simpy.resources.store import StorePut, StoreGet

import eflips

# Settings for dates on x axis
abs_time_fmt = "%"
base_date = datetime(2018, 12, 4, 0, 0)  # arbitrary Tuesday 0:00
datefmt_general = "%a %H:%M"  # format to "Mon 00:00"
datefmt_major = "%a"  # 'Mon'
datefmt_major2 = "%H:%M\n%a"  # '00:00\nMon'
datefmt_minor = "%H:%M"  # '00:00'
xdatespacing_major = 86400  # show major x ticks at these multiplier seconds
# divisor for of xdatespacing_major to show  minor ticks at (e.g. 4 for every
# 6 hours if xdatespacing_major is 86400)
minor_intervals_per_major_tick = 4

color_bvg_yellow = "#f0d722"


class ArrivalLog:
    """Container for data to be logged when a vehicle arrives at a depot.
    Independent from the DataLogger. Temporary, may be replaced by a DataLogger
    rework.
    """

    def __init__(self, t, vehicle):
        self.arrival_time = t
        self.vehicle = vehicle
        self.trip = vehicle.trip
        self.energy = vehicle.battery.energy
        self.energy_real = vehicle.battery.energy_real


class BatteryLog:
    """Container for logging data related to charging. Independent from the
    DataLogger. Temporary, may be replaced by a DataLogger rework.

    Parameters:
    event_name: [str]
    """

    def __init__(self, t, vehicle, event_name):
        self.t = t
        self.energy = vehicle.battery.energy
        self.energy_real = vehicle.battery.energy_real
        self.event_name = event_name


class DepotEvaluation:
    """Container for data of one depot simulation run to be accessible after
    the simulation. Provides tools for evaluating and exporting data.

    Parameters:
    depotsim: [DepotSimulation] instance

    Attributes:
    results: [dict] storing results of evaluation after the simulation
    xlim: [tuple] of x axis limits for some plots and calculations. Default is
        (0, self.SIM_TIME)
    arrival_logs: [list] of ArrivalLog objects
    sl_logs: [dict] with logged stress level values (by self.calculate_sl)
    cm_report: [DepotAnalysis]

    """

    def __init__(self, depotsim):
        self.depotsim = depotsim

        self.env = depotsim.env
        self.depot = None
        self.configurator = None
        self.timetable = None
        self.vehicle_generator = None
        self.gc = globalConstants
        self.results = {}
        self.path_results = globalConstants["depot"]["path_results"]

        self.SIM_TIME = None
        self.xlim = None
        self.sim_start_datetime = None

        self.power_logs = {0: 0}
        self.arrival_logs = []
        self.sl_logs = OrderedDict()

        self.cm_report = None

        # Further preparation of results
        self.results["idle_time"] = {}

    def complete(self):
        """Complete the initialization by creating references that were not
        accessible during init. Must be called before simulation start.
        """
        self.depot = self.depotsim.depot
        self.configurator = self.depotsim.configurator
        self.cm_report = DepotAnalysis(self.depotsim)
        self.depot.evaluation = self
        self.timetable = self.depotsim.simulation_host.timetable
        self.vehicle_generator = self.depotsim.simulation_host.vg

        self.path_results = globalConstants["depot"]["path_results"]
        self.SIM_TIME = self.gc["general"]["SIMULATION_TIME"]

        if "SIMULATION_START_DATETIME" in self.gc["general"]:
            self.sim_start_datetime = self.gc["general"]["SIMULATION_START_DATETIME"]

        self.xlim = (0, self.SIM_TIME)

    @property
    def now_repr(self):
        """Return the current system date and time as formatted string."""
        return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    def log_arrival(self, vehicle):
        self.arrival_logs.append(ArrivalLog(self.env.now, vehicle))

    @property
    def current_sl(self):
        """The stress level [int] describes the current planning situation of
        dispatch for one vehicle type or vehicle type group.
        *sl_period* in globalConstants defines how far in the
        future from now trips and vehicles are considered.
        stress level = - number of trips that have no vehicle assigned
                + number of vehicles that have no trip assigned
        Both summands are calculated for now + sl_period.
        One of or both summands for the same vehicle type or type group are
        always 0 since they describe excesses that are matched against each
        other.

        Return only the data required for calculating sl to reduce the effort
        during the simulation.
        """
        sl_period = self.env.now + globalConstants["depot"]["sl_period"]

        # Get trips that are supposed to depart from now until now + sl_period
        trips = [
            trip
            for trip in self.depot.pending_departures
            if trip.vehicle is None and trip.std <= sl_period
        ]

        # Find vehicles that are ready to depart or will be in now + sl_period
        # and match them with trips from above, considering vehicle types.
        # c_vehicles will contain the number of remaining unmatched vehicles by
        # vehicle type
        c_vehicles = Counter()
        for area in self.depot.depot_control.departure_areas.stores:
            for vehicle in area.vehicles:
                if vehicle.trip is None:
                    etc = vehicle.dwd.etc_processes
                    if (
                        etc is eflips.depot.EstimateValue.COMPLETED
                        or isinstance(etc, int)
                        and etc <= sl_period
                    ):
                        # flexprint('Considering vehicle %s. Number of trips: %d' % (vehicle.ID, len(trips)),
                        #           env=self.env)
                        match = False
                        for trip in trips:
                            if vehicle.vehicle_type in trip.vehicle_types:
                                trips.remove(trip)
                                match = True
                                # flexprint('Found match for vehicle %s. trips left: %d' % (vehicle.ID, len(trips)), env=self.env)
                                break
                        if not match:
                            if vehicle.vehicle_type.group is not None:
                                k = vehicle.vehicle_type.group.vehicle_types_joinedstr
                            else:
                                k = vehicle.vehicle_type.ID
                            c_vehicles[k] += 1
                            # flexprint("Didn't find match for vehicle %s. Number of trips: %d" % (vehicle.ID, len(trips)),
                            #     env=self.env)

        # Count the remaining unmatched trips by vehicle type
        c_trips = Counter()
        for trip in trips:
            c_trips[trip.vehicle_types_joinedstr] += 1

        return c_trips, c_vehicles

    def log_sl(self, *args, **kwargs):
        if globalConstants["depot"]["log_sl"]:
            trips, vehicles = self.current_sl
            self.sl_logs[self.env.now] = {"trips": trips, "vehicles": vehicles}

    def excel_report_full(self):
        report = Report(self.depotsim)
        report.export_to_excel()

    # Evaluation and plots for the whole depot
    def nvehicles_initstore(
        self,
        show=True,
        save=False,
        basefilename="nvehicles_initstore",
        formats=("png",),
    ):
        """Vehicle count over time in the depot's init_store. Plot as line."""

        plot_series = self.depot.init_store.logger.get_valList("count", self.SIM_TIME)

        plot_series = to_prev_values(plot_series)

        if show or save:
            fig, ax = baseplot(show)

            ax.plot(plot_series)
            ax.set_title(
                'Depot "' + self.depot.ID + '" number of vehicles in init store'
            )
            plt.xlabel("Time [s]")
            plt.ylabel("No of vehicles")
            to_dateaxis(ax)
            ax.set_ylim(bottom=0)

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def nvehicles_initstore_end(
        self,
        show=True,
        save=False,
        basefilename="nvehicles_initstore_end",
        formats=("png",),
    ):
        """Vehicle count per type in the depot's init_store at the end of
        the simulation. Plot in a histogram.
        """
        occurrences = [
            vehicle.vehicle_type.ID for vehicle in self.depot.init_store.items
        ]
        if not occurrences:
            print("VehicleStoreInit is empty at the end of the simulation.")
        else:
            vehicle_types = list(set(occurrences))
            typemap = {}
            for vehicletypeNo, vehicletype in enumerate(vehicle_types):
                typemap[vehicletype] = vehicletypeNo

            plot_series = [typemap[vt] for vt in occurrences]

            counter = Counter(occurrences)
            print("Occurences: %s" % dict(counter))

            if show or save:
                fig, ax = baseplot(show)

                plt.hist(plot_series, bins=len(vehicle_types))
                plt.xlabel("Vehicle Type")
                plt.ylabel("Count")
                plt.title("Vehicles in init_store at the end of the simulation")

                nxticks = range(len(vehicle_types))
                plt.xticks(nxticks, vehicle_types)

                adjust_plt_hist()

                if show:
                    fig.show()
                if save:
                    filename = self.path_results + basefilename
                    savefig(fig, filename, formats)
                if not show:
                    plt.close(fig)

    def nvehicles_used_calculation(self):
        """Number of vehicles that were used by type"""
        occurrences = Counter()
        for vehicle in self.vehicle_generator.items:
            if vehicle.system_entry:
                occurrences[vehicle.vehicle_type.ID] += 1

        return occurrences

    def nvehicles_used(
        self,
        show=True,
        save=False,
        basefilename="nvehicles_required",
        formats=("png",),
        language=("eng"),
    ):
        """Number of vehicles that were used by type. Plot in a bar chart."""

        occurrences = self.nvehicles_used_calculation()
        print("Occurences: %s" % occurrences)
        print("Total: %d" % sum(occurrences.values()))

        if show or save:
            fig, ax = baseplot(show)
            setting_language(language)

            labels, values = zip(*occurrences.items())

            indexes = np.arange(len(labels))
            width = 0.6

            bars = plt.bar(indexes, values, width)
            plt.xticks(indexes, labels)

            # Attach a text label above each bar, displaying its height
            for bar in bars:
                height = bar.get_height()
                ax.annotate(
                    "{}".format(height),
                    xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3),  # 3 points vertical offset
                    textcoords="offset points",
                    ha="center",
                    va="bottom",
                )

            if language == "eng":
                plt.xlabel("Vehicle Type")
            else:
                plt.xlabel("Fahrzeugtyp")

            if language == "eng":
                plt.ylabel("Amount")
            else:
                plt.ylabel("Anzahl")

            plt.title("Number of vehicles that were used by type")

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def nvehicles_total(
        self, show=True, save=False, basefilename="nvehicles_total", formats=("png",)
    ):
        """Total vehicle count over time in the depot in total. Plot as line."""
        # Sum up the counts of all areas in the depot
        plot_series = np.zeros(self.SIM_TIME, dtype=np.int32)
        for area in self.depot.list_areas:
            y = area.logger.get_valList("count", self.SIM_TIME)
            y = to_prev_values(y)
            y = np.array(y)
            plot_series += y

        if len(plot_series) > 86400:
            print("After the first day:")
            print("\tMin: %d" % np.min(plot_series[86400:]))
            print("\tMax: %d" % np.max(plot_series[86400:]))

        if show or save:
            fig, ax = baseplot(show)

            ax.plot(plot_series)

            # # engl. labels
            ax.set_title('Depot "' + self.depot.ID + '" total number of vehicles')
            plt.ylabel("No of vehicles")

            # german labels
            # ax.set_title('Tagesganglinie Depot')
            # plt.ylabel('Anzahl Fahrzeuge')

            plt.xlim(*self.xlim)
            # plt.ylim(0, 141)
            to_dateaxis(ax)

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def arrival_soc(
        self,
        typs={},
        show=True,
        save=False,
        basefilename="arrival_soc",
        formats=("png",),
        language=("eng"),
    ):
        """Plot the SoC of vehicles upon arrival at the depot."""

        soc_values = pd.DataFrame(columns=["time", "soc", "color"])
        default_color_used = False
        for arrival_log in self.depot.evaluation.arrival_logs:
            if self.xlim[0] <= arrival_log.arrival_time <= self.xlim[1]:
                soc = arrival_log.energy / arrival_log.energy_real
                if arrival_log.vehicle.vehicle_type.ID in typs.keys():
                    color = typs[arrival_log.vehicle.vehicle_type.ID]
                else:
                    default_color_used = True
                    color = "#009ACD"  # light blue hard coded later one, so if you change here also change below
                soc_values.loc[arrival_log.arrival_time] = [
                    arrival_log.arrival_time,
                    soc,
                    color,
                ]

        count_negative_soc = 0
        count_negativity_soc = 0
        for index, row in soc_values.iterrows():
            if row.soc < 0:
                count_negative_soc += 1
            if row.soc == globalConstants["depot"]["reset_negative_soc_to"]:
                count_negativity_soc += 1
        share_negative_soc = count_negative_soc / soc_values.shape[0] * 100
        share_negativity_soc = count_negativity_soc / soc_values.shape[0] * 100
        print(
            "Amount of arrivals with negative SoC: %s = %s %% of %s "
            % (count_negative_soc, share_negative_soc, soc_values.shape[0])
            + "values in total."
        )
        print(
            "Amount of arrivals with SoC value equal to the reset value %s: "
            "%s = %s %% of %s "
            % (
                globalConstants["depot"]["reset_negative_soc_to"],
                count_negativity_soc,
                share_negativity_soc,
                soc_values.shape[0],
            )
            + "values in total."
        )

        if show or save:
            fig, ax = baseplot(show)
            setting_language(language)

            soc_values.plot.scatter(
                x="time", y="soc", s=3, c=list(soc_values.color), ax=ax
            )
            if language == "eng":
                plt.xlabel("Time")
            else:
                plt.xlabel("Zeit")
            plt.ylabel("SoC")
            plt.title("SoC values upon arrival at the depot")
            plt.ylim(0, 1)
            ax.yaxis.grid(True)
            patches = []
            for key, value in typs.items():
                patches.append(mpatches.Patch(color=value, label=key))
            if default_color_used:
                patches.append(mpatches.Patch(color="#009ACD", label="all other"))
            plt.legend(handles=patches)
            to_dateaxis(ax)

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def departure_soc(
        self,
        typs={},
        show=True,
        save=False,
        basefilename="departure_soc",
        formats=("png",),
        language=("eng"),
    ):
        """Plot the SoC of vehicles upon departure from the depot."""
        soc_values = pd.DataFrame(columns=["time", "soc", "color"])
        default_color_used = False

        for vehicle in self.vehicle_generator.items:
            for log in vehicle.battery_logs:
                if log.event_name == "consume_start":
                    if self.xlim[0] <= log.t <= self.xlim[1]:
                        soc = log.energy / log.energy_real
                        if vehicle.vehicle_type.ID in typs.keys():
                            color = typs[vehicle.vehicle_type.ID]
                        else:
                            default_color_used = True
                            color = "#009ACD"  # light blue hard coded later one, so if  you change here also change below
                        soc_values.loc[log.t] = [log.t, soc, color]

        print("\n Summary of departure SoC: ")
        print("min: ", round(soc_values.soc.min(), 3))
        print("max: ", round(soc_values.soc.max(), 3))
        print("mean: %.2f" % round(soc_values.soc.mean(), 3))
        print("median: %.2f" % round(soc_values.soc.median(), 3))

        print("number of departures: ", soc_values.shape[0])

        if show or save:
            fig, ax = baseplot(show)
            setting_language(language)

            soc_values.plot.scatter(
                x="time", y="soc", s=3, c=list(soc_values.color), ax=ax
            )
            if language == "eng":
                plt.xlabel("Time")
            else:
                plt.xlabel("Zeit")
            plt.ylabel("SoC")
            plt.title("SoC values upon departure from the depot")
            ax.set_axisbelow(True)
            ax.yaxis.grid(True)

            patches = []
            for key, value in typs.items():
                patches.append(mpatches.Patch(color=value, label=key))
            if default_color_used:
                patches.append(mpatches.Patch(color="#009ACD", label="all other"))
            plt.legend(handles=patches)
            to_dateaxis(ax)

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def get_periods(self, vehicle, vehicle_no, periods):
        """Helper function for self.vehicle_periods."""
        y = vehicle_no + 1
        nonprocs = ["depot general", "park"]
        proc_IDs = []
        for k in periods:
            if k not in nonprocs:
                proc_IDs.append(k)

        data = {
            "depot general": {"xranges": [], "yranges": (y, 0.95), "triptexts": []},
            "park": {"xranges": [], "yranges": (y, 0.75), "slottexts": []},
        }

        # Get parking time and total depot time
        logged_data_areas = vehicle.logger.loggedData["dwd.current_area"]
        logged_data_slots = vehicle.logger.loggedData["dwd.current_slot"]
        current = None
        previous = None
        start_depot = None
        start_park = None
        depot_end_found = True
        park_end_found = True
        first_period_found = False

        # Determine depot and parking periods and trip info
        for t in logged_data_areas:
            current = logged_data_areas[t]
            if (
                current == previous
            ):  # skip steplog entries where nothing changed for this vehicle
                continue
            elif current is not None:
                if previous is None:  # enter depot
                    start_depot = t
                    assert depot_end_found
                    depot_end_found = False

                    # arrival trip info (only first trip for each vehicle)
                    if not first_period_found:
                        trip = next(
                            (ti for ti in vehicle.finished_trips if t == ti.ata), None
                        )

                        # TODO there is no clear message on which area is too small
                        assert trip is not None, (
                            "Possible reason: vehicle has a finished trip, "
                            "but never entered the depot because there was "
                            "not enough space (e.g. arrival area is too small)"
                        )

                        if self.xlim[0] <= t <= self.xlim[1]:
                            data["depot general"]["triptexts"].append(
                                {
                                    "ID": trip.ID,
                                    "line_name": trip.line_name,
                                    "x": t,
                                    "y": y + 0.5,
                                    "delay": trip.ata - trip.sta,
                                    "as_arrival": True,
                                }
                            )

                    first_period_found = True

                if current.issink:  # enter parking lot
                    start_park = t
                    assert park_end_found
                    park_end_found = False

            elif current is None:  # exit depot
                depot_end_found = True
                park_end_found = True
                data["depot general"]["xranges"].append((start_depot, t - start_depot))
                data["park"]["xranges"].append((start_park, t - start_park))

                if self.xlim[0] <= start_park <= self.xlim[1]:
                    data["park"]["slottexts"].append(
                        {
                            "slot_no": previous.ID + str(logged_data_slots[start_park]),
                            "x": start_park,
                            "y": y + 0.5,
                        }
                    )

                # departure trip info
                trip = next((ti for ti in vehicle.finished_trips if t == ti.atd), None)
                if trip is None:
                    # trip wasn't finished until sim end, it's the current
                    trip = vehicle.trip
                assert trip is not None

                if self.xlim[0] <= t <= self.xlim[1]:
                    data["depot general"]["triptexts"].append(
                        {
                            # 'ID': trip.ID.split('_')[0],
                            "ID": trip.ID,
                            "line_name": trip.line_name,
                            "x": t,
                            "y": y + 0.5,
                            "delay": trip.atd - trip.std,
                            "as_arrival": False,
                        }
                    )

            previous = current

        # Stay lasted until sim time end, complete and append final entry
        if not depot_end_found:
            data["depot general"]["xranges"].append(
                (start_depot, self.SIM_TIME - start_depot)
            )
        if not park_end_found:
            data["park"]["xranges"].append((start_park, self.SIM_TIME - start_park))

            if self.xlim[0] <= start_park <= self.xlim[1]:
                data["park"]["slottexts"].append(
                    {
                        "slot_no": current.ID + str(logged_data_slots[start_park]),
                        "x": start_park,
                        "y": y + 0.5,
                    }
                )

        # Preliminary until proper switch is implemented
        if "depot general" not in periods:
            del data["depot general"]
        if "park" not in periods:
            del data["park"]

        # Determine process periods
        if proc_IDs:
            for procID in proc_IDs:
                data[procID] = {"xranges": [], "yranges": (y, 0.5)}
            logged_data_processes = vehicle.logger.loggedData[
                "dwd.active_processes_copy"
            ]
            for t in logged_data_processes:
                for proc in logged_data_processes[t]:
                    if proc.ID in data:
                        # Fill xranges with tuples (start, duration)
                        for startno, start in enumerate(proc.starts):
                            if startno != len(proc.ends):
                                # startno has a match in end
                                data[proc.ID]["xranges"].append(
                                    (start, proc.ends[startno] - start)
                                )
                            else:
                                # ends is shorter than starts; proc lasted
                                # until sim time end
                                data[proc.ID]["xranges"].append(
                                    (start, self.SIM_TIME - start)
                                )

        data["trip"] = {"xranges": [], "yranges": (y, 0.75), "triptexts": []}

        # Read trip data from vehicle
        for trip in vehicle.finished_trips:
            data["trip"]["xranges"].append((trip.atd, trip.duration))

        return data

    def vehicle_periods(
        self,
        periods=None,
        show=True,
        save=False,
        basefilename="vehicle_periods",
        formats=("pdf",),
        show_vehicle_count=True,
        vehicle_count_color="#1f77b4",
        show_total_power=True,
        total_power_color="#d62728",
        show_annotates=True,
        vehicle_ID="all",
        smart_charging=None,
        smart_charging_color="forestgreen",
    ):
        """Plot the period history of all vehicles in a horizontal broken bar
        chart with one y tick per vehicle.


        periods: [dict] of process name and color as key-value pairs.
        vehicle_ID: [str] Possibility to filter periods to plot by vehicle ID. Does NOT work with smart charging.
            Examples:
                'all': show all vehicles
                'EN_DC 20': show only the specific vehicle with ID 'EN_DC 20'
                'EN': show all vehicles with IDs containing 'EN'
        smart_charging: [SmartCharging] if an proper object is given, the charging data from smart_charging will be ploted

        See eflips.depot.plots for an example call of this function.
        """
        if periods is None:
            periods = {"depot general": "yellow"}

        # Set sim time lower bound to avoid inaccuracy of first day
        # TODO make sure the correct xmin_calcs for plotting
        xmin_calcs = 0
        # xmin_calcs = 86400 if self.SIM_TIME > 86400 else 0

        # Get data for periods
        vehicledata = {}
        self.results["vehicle_periods"] = {}
        self.results["vehicle_periods"]["vehicledata"] = vehicledata
        vehicle_no = 0
        for vehicle in self.vehicle_generator.items:
            if vehicle_ID == "all" or vehicle_ID in vehicle.ID:
                vehicledata[vehicle.ID] = {}
                vehicledata[vehicle.ID]["plotdata"] = self.get_periods(
                    vehicle, vehicle_no, periods
                )
                vehicle_no += 1
        if not vehicledata:
            print(
                "No data for vehicles found. Possibly argument 'vehicle_ID' "
                "doesn't match any ID."
            )
            return

        if isinstance(smart_charging, eflips.depot.smart_charging.SmartCharging):
            for (
                vehicle_ID,
                data,
            ) in smart_charging.export_for_vehicle_periods_smart().items():
                vehicledata[vehicle_ID]["plotdata"]["smart_charging"] = data

            periods["smart_charging"] = smart_charging_color

        # Print and plot periods
        if show or save:
            # Period history plot
            axislabelsize = 10
            ticklabelsize = 8
            legendfontsize = 8

            fig, ax1 = plt.subplots()

            ax1.set_xlim(*self.xlim)

            plot_title = "Period History of Vehicles"
            plot_title += " - " + self.configurator.templatename_display

            plt.title(plot_title)
            # ax1.xaxis.grid(True)
            # ax1.yaxis.grid(True)
            # statistics
            dwell_time_total = 0
            dwell_time_counter = 0

            for ID in vehicledata:
                print("Vehicle " + ID + ":")
                print("\tperiods:")
                for periodID in vehicledata[ID]["plotdata"]:
                    print("\t", periodID, ": ")
                    print(
                        "\t\txranges: ",
                        vehicledata[ID]["plotdata"][periodID]["xranges"],
                    )
                    print(
                        "\t\tyranges: ",
                        vehicledata[ID]["plotdata"][periodID]["yranges"],
                    )
                    # if periodID == 'depot general':
                    #     for start, duration in vehicledata[ID]['plotdata'][periodID]['xranges']:
                    #         dwell_time_total += duration
                    #         dwell_time_counter += 1
                    #         print(dwell_time_total, duration, dwell_time_counter)
                    facecolor = periods[periodID]
                    ax1.broken_barh(
                        vehicledata[ID]["plotdata"][periodID]["xranges"],
                        vehicledata[ID]["plotdata"][periodID]["yranges"],
                        facecolors=facecolor,
                    )
            ylim_upper = len(vehicledata) + 1.5
            ax1.set_ylim(0.5, ylim_upper)
            ax1.set_yticks(
                [(vehicle_no + 1) + 0.5 for vehicle_no in range(len(vehicledata))]
            )
            # ax1.set_yticklabels(vehicledata.keys(), fontsize=ticklabelsize)
            ax1.set_yticklabels(
                [i + 1 if (i + 1) % 10 == 0 else "" for i in range(len(vehicledata))],
                fontsize=ticklabelsize,
            )
            ax1.set_ylabel("Vehicle ID", fontsize=axislabelsize)
            ax1.set_zorder(-3)
            # ax1.set_axisbelow(True)

            # Create variables for the legend
            marks = [mpatch.Rectangle((0, 0), 1, 1, fc=clr) for clr in periods.values()]
            names = list(periods.keys())

            # Additional plot in same window: vehicle count
            # Sum up the counts of all areas in the depot
            if show_vehicle_count:
                y2 = np.zeros(self.SIM_TIME, dtype=np.int32)
                for area in self.depot.list_areas:
                    nv = area.logger.get_valList("count", SIM_TIME=self.SIM_TIME)
                    nv = to_prev_values(nv)
                    nv = np.array(nv)
                    y2 += nv

                # Plot on own y axis that aligns with periods axis
                # Correct y and axis to align with period history middle-placed
                # y ticks
                ax2 = ax1.twinx()
                y2_offset = y2 + 0.5  # offset of half the period bar height
                count_plot = ax2.plot(y2_offset, color=vehicle_count_color)

                ax2.set_ylim(0.5, len(vehicledata) + 1.5)
                ax2.set_yticks(
                    [(vehicle_no) + 0.5 for vehicle_no in range(len(vehicledata) + 1)]
                )
                ax2.set_yticklabels(range(len(vehicledata) + 1), fontsize=ticklabelsize)
                ax2.set_zorder(-2)
                # Hide the own y axis (unhide if vehicle IDs are not numbers)
                ax2.get_yaxis().set_visible(False)
                # ax2.set_ylabel('Vehicle count', fontsize=axislabelsize)

                # Extend legend variables
                marks.append(count_plot[0])
                names.append("vehicle count")

                print("\nMin vehicle count after first day: ", min(y2[xmin_calcs:]))
                print("Max vehicle count after first day: ", max(y2[xmin_calcs:]))

            # Additional plot in same window: total power
            if show_total_power:
                seriesname = "Power [kW]"
                if isinstance(
                    smart_charging, eflips.depot.smart_charging.SmartCharging
                ):
                    power = smart_charging.power.used_power_smart.sort_index()
                    power = power.to_dict()
                    y3 = discrete2continuous_logs(power, first=0)
                else:
                    y3 = discrete2continuous_logs(
                        self.depot.evaluation.power_logs, first=0
                    )
                ax3 = ax1.twinx()
                ax3.set_ylabel(seriesname, fontsize=axislabelsize)
                power_plot = ax3.plot(y3, color=total_power_color)
                ax3.set_zorder(-1)

                max_possible_power = sum(
                    [
                        ci.max_power
                        for ci in self.depot.resources.values()
                        if isinstance(ci, eflips.depot.DepotChargingInterface)
                    ]
                )
                total_max_power = max(y3[xmin_calcs:]) + max(y3[xmin_calcs:]) * 0.1
                ax3.set_ylim(top=total_max_power)
                ax3.yaxis.set_major_locator(ticker.MultipleLocator(250))

                align_yaxis(ax1, 0.5, ax3, 0)
                ax3.tick_params(axis="y", which="both", labelsize=ticklabelsize)

                # Extend legend variables
                marks.append(power_plot[0])
                names.append("power")

                print("\nMin power after first day: ", min(y3[xmin_calcs:]))
                print("Max power after first day: ", max(y3[xmin_calcs:]))

            # Extension of periods: Annotate depot stay periods with trip IDs
            if show_annotates:
                ax4 = ax1.twinx()  # create new axes for separate zorder
                ax4.set_zorder(0)
                ax4.set_ylim(0.5, ylim_upper)
                ax4.get_yaxis().set_visible(False)
                font = FontProperties(size=1)
                alignment = {
                    "horizontalalignment": "center",
                    "verticalalignment": "center",
                }
                rotation = 0  # degrees
                xoffset_arr = -1500  # seconds
                xoffset_dep = 1500  # seconds
                delay_color = "firebrick"

                for vID in vehicledata:
                    for tripInfo in vehicledata[vID]["plotdata"]["depot general"][
                        "triptexts"
                    ]:
                        xoffset = xoffset_arr if tripInfo["as_arrival"] else xoffset_dep
                        color_txt = delay_color if tripInfo["delay"] != 0 else "black"
                        text = ax4.text(
                            tripInfo["x"] + xoffset,
                            tripInfo["y"],
                            tripInfo["line_name"],
                            fontproperties=font,
                            **alignment,
                            color=color_txt,
                            rotation=rotation,
                        )

                        x_pos = text.get_position()[0]
                        # set texts outside xlim to invisible
                        if not (self.xlim[0] < x_pos < self.xlim[1]):
                            text.set_visible(False)

                        # Highlight the delay
                        if tripInfo["delay"] != 0:
                            circle = mpatch.Ellipse(
                                (tripInfo["x"] + xoffset - 500, tripInfo["y"]),
                                # 10000, 1.15,
                                1500,
                                1.643,
                                color="#eb9999",
                            )
                            ax1.add_artist(circle)  # ax1 to ensure background
                            ax4.text(
                                tripInfo["x"] + xoffset,
                                tripInfo["y"] - 0.7,
                                str(int(tripInfo["delay"] / 60)) + " min",
                                fontproperties=font,
                                **alignment,
                                color=color_txt,
                                rotation=rotation,
                            )

                # Extension of periods: Annotate parking periods with slot numbers
                font = FontProperties(size=1)
                alignment = {
                    "horizontalalignment": "left",
                    "verticalalignment": "bottom",
                }
                rotation = 0  # degrees
                xoffset = 500  # seconds
                yoffset = -0.495  # y ticks
                color_txt = "#474747"

                for vID in vehicledata:
                    for slotInfo in vehicledata[vID]["plotdata"]["park"]["slottexts"]:
                        text = ax4.text(
                            slotInfo["x"] + xoffset,
                            slotInfo["y"] + yoffset,
                            slotInfo["slot_no"],
                            fontproperties=font,
                            color=color_txt,
                            rotation=rotation,
                        )
                        x_pos = text.get_position()[0]
                        # set texts outside xlim to invisible
                        if not (self.xlim[0] < x_pos < self.xlim[1]):
                            text.set_visible(False)

            # Convert x axis seconds to dates
            to_dateaxis(ax1)
            plt.xlim(*self.xlim)

            # Create one legend for all plots
            legend = plt.legend(marks, names, loc="upper left", fontsize=legendfontsize)
            legend.get_frame().set_alpha(1)
            # legend.set_zorder(50)

            if show:
                plt.show()

            if save:
                # Set resolution to a default
                dpi = fig.get_dpi()
                fig.set_size_inches(1920.0 / float(dpi), 948.0 / float(dpi))
                filename = os.path.join(self.path_results, basefilename)
                savefig(fig, filename, formats=formats, dpi="figure")
            if not show:
                plt.close(fig)

    def idle_time_dist(
        self,
        charge_IDs=("charge",),
        vehicle_types="all",
        show=True,
        save=False,
        basefilename="idle_time_dist",
        formats=("png",),
        bins=100,
        color="#1f77b4",
    ):
        """Length of parking time after the end of charging until departure.
        Data of the first simulation day is excluded.

        charge_IDs: [tuple] IDs of the charge processes
        vehicle_types: The vehicle types to calculate the idle time for. Must
            be 'all' [str] to include all types or an iterable of vehicle type
            IDs as str (e.g. ['EN', 'GN']).
        """
        # Check input
        for charge_ID in charge_IDs:
            if charge_ID not in self.depot.processes.keys():
                print('Cannot find process ID "%s".' % charge_ID)
                return
        if vehicle_types != "all":
            all_types = list(self.gc["depot"]["vehicle_types"].keys())
            for vt in vehicle_types:
                if vt not in all_types:
                    print('Cannot find vehicle type "%s".' % vt)
                    return

        # Get the data
        periods = {"park": "burlywood"}  # arbitrary color
        for charge_ID in charge_IDs:
            periods[charge_ID] = "#2ca02c"  # arbitrary color
        self.vehicle_periods(periods=periods, show=False, save=False)
        vehicledata = self.results["vehicle_periods"]["vehicledata"]
        # Set sim time lower bound to avoid inaccuracy of first day
        xmin_calcs = 86400 if self.SIM_TIME > 86400 else 0

        idle_times = []
        for vehicleID in vehicledata:
            if (
                vehicle_types == "all"
                or self.vehicle_generator.select(vehicleID).vehicle_type.ID
                in vehicle_types
            ):
                for charge_ID in charge_IDs:
                    pp = vehicledata[vehicleID]["plotdata"]["park"]["xranges"]
                    # Get the ID of the correct charging process
                    # c_ID = next(charge_ID for charge_ID in charge_IDs
                    # if charge_ID in vehicledata[vehicleID]['plotdata'].keys())
                    c_ID = charge_ID
                    cc = vehicledata[vehicleID]["plotdata"][c_ID]["xranges"]
                    for p in pp:
                        if p[0] >= xmin_calcs:
                            # Find corresponding charging period, if there is one
                            # Ignore charging periods that lasted until sim time end
                            c = next(
                                (
                                    ci
                                    for ci in cc
                                    if p[0] <= ci[0] <= p[0] + p[1]
                                    and ci[0] + ci[1] != self.SIM_TIME
                                ),
                                None,
                            )
                            if c:
                                idle_times.append((p[0] + p[1] - (c[0] + c[1])) / 60)
        # idle_times.append(0)
        # idle_times.append(900)

        # Calculate, save and print additional figures
        if vehicle_types == "all":
            vt_joined = vehicle_types
        elif len(vehicle_types) == 1:
            vt_joined = vehicle_types[0]
        else:
            vt_joined = ", ".join(vehicle_types)

        self.results["idle_time"][vt_joined] = {}
        idle_times_array = np.asarray(idle_times)
        print("\nIdle times after first day for vehicle types: %s" % vt_joined)
        it_count = len(idle_times_array)
        print("\tcount: ", it_count)
        it_min = idle_times_array.min()
        self.results["idle_time"][vt_joined]["min"] = it_min
        print("\tmin: ", int(it_min * 60), " s; = ", "%.2f" % it_min, " min")
        it_max = idle_times_array.max()
        self.results["idle_time"][vt_joined]["max"] = it_max
        print("\tmax: ", int(it_max * 60), " s; = ", "%.2f" % it_max, " min")
        it_mean = np.mean(idle_times_array)
        self.results["idle_time"][vt_joined]["average"] = it_mean
        print("\taverage: ", int(it_mean * 60), " s; = ", "%.2f" % it_mean, " min")
        it_median = np.median(idle_times_array)
        print("\tmedian: ", int(it_median * 60), " s; = ", "%.2f" % it_median, " min")
        print("\toccurrences of 0: ", np.count_nonzero(idle_times_array == 0))
        print("\toccurrences of <= 15 min: ", np.count_nonzero(idle_times_array <= 15))

        if show or save:
            fig, ax = baseplot(show)

            plt.hist(idle_times, bins=bins, range=[0, 900], color=color)
            plt.xlabel("Time [min]")
            plt.xticks(np.arange(0, 960, step=60))
            plt.ylabel("Number of occurrences")
            plt.title(
                "Distribution of idle time (parking time after "
                "charging) for vehicle types: %s" % vt_joined
            )
            plt.xlim(left=0)
            plt.ylim(bottom=0)
            plt.grid()

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def total_power(
        self, show=True, save=False, basefilename="total_power", formats=("png",)
    ):
        """Plot the total power in the depot over time as line."""
        plot_title = "Depot total power"

        y = discrete2continuous_logs(self.power_logs, first=0)
        print("Min power: ", min(y))
        print("Max power: ", max(y))

        if show or save:
            fig, ax = baseplot(show)

            ax.plot(y)
            plt.title(plot_title)
            plt.ylabel("Power [kW]")
            plt.xlim(*self.xlim)
            to_dateaxis(ax)
            plt.grid()

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def ci_power(
        self,
        cis=[1, 1],
        show=True,
        save=False,
        basefilename="total_power",
        formats=("png",),
        language=("eng"),
    ):
        """Plot the total power over time as line of charging infrastructures (cis)."""
        # Something wrong with discrete2continuous_logs!! Looks fine for EN, but not for HPC

        plot_title = "Total power of ci_" + str(cis[0]) + " - ci_" + str(cis[1])

        power_logs_cis = np.empty_like(
            discrete2continuous_logs(
                self.depot.resources["ci_" + str(cis[0])].power_logs_ci, first=0
            )
        )
        ci_counter = 0

        for ci_no in range(cis[0], cis[1] + 1):
            power_logs_ci = discrete2continuous_logs(
                self.depot.resources["ci_" + str(ci_no)].power_logs_ci, first=0
            )
            # print(power_logs_ci)
            power_logs_cis = np.add(power_logs_ci, power_logs_cis)
            print("ci", ci_no, ": ", str(max(power_logs_ci)))
            if max(power_logs_ci) > 0:
                ci_counter += 1
        print("Amount used cis: ", ci_counter)
        print(power_logs_cis, len(power_logs_cis))
        y = power_logs_cis
        print("Min power: ", min(y))
        print("Max power: ", max(y))

        if show or save:
            fig, ax = baseplot(show)
            setting_language(language)

            if language == "eng":
                plt.xlabel("Time")
            else:
                plt.xlabel("Zeit")

            if language == "eng":
                plt.ylabel("Power [kW]")
            else:
                plt.ylabel("Leistung [kW]")

            ax.plot(y)
            plt.title(plot_title)
            plt.xlim(*self.xlim)
            to_dateaxis(ax)
            plt.grid()

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def calculate_sl(self):
        """Calculation of stress level (sl) values after the simulation."""
        for log in self.sl_logs.values():
            log["sl"] = self.calculate_sl_single(log)
        self.sl_logs.calculated = True

    @staticmethod
    def calculate_sl_single(log):
        """Calculate sl for a single log."""
        sl = log["trips"].copy()
        for k in sl:
            sl[k] *= -1
        sl.update(log["vehicles"])
        return sl

    def sl_single(
        self,
        vehicle_type_ID,
        show=True,
        save=False,
        basefilename="stress_level",
        formats=("pdf",),
    ):
        """Plot the stress level of *vehicle_type_ID* and its calculation
        parameters over time as line.
        """
        if not self.gc["depot"]["log_sl"]:
            raise ValueError(
                "stress level (sl) values were not logged. Rerun the simulation "
                "with globalConstants['depot']['log_sl'] = True."
            )
        vt_obj = next(
            (
                vt
                for vt in self.gc["depot"]["vehicle_types_obj"]
                if vt.ID == vehicle_type_ID
            ),
            None,
        )
        if vt_obj is None:
            raise ValueError("Unknown vehicle type '%s'" % vehicle_type_ID)
        if vt_obj.group is not None:
            vehicle_type_ID = ", ".join(vt.ID for vt in vt_obj.group.types)

        if not hasattr(self.sl_logs, "calculated"):
            self.calculate_sl()

        sl = discrete2continuous_logs(
            self.sl_logs, first=0, subkeys=["sl", vehicle_type_ID]
        )
        n_trips = discrete2continuous_logs(
            self.sl_logs, first=0, subkeys=["trips", vehicle_type_ID]
        )
        n_vehicles = discrete2continuous_logs(
            self.sl_logs, first=0, subkeys=["vehicles", vehicle_type_ID]
        )

        # Set sim time lower bound to avoid inaccuracy of first day
        xmin_calcs = 86400 if self.SIM_TIME > 86400 else 0

        # Print results
        period = globalConstants["depot"]["sl_period"]
        print(
            "Values for %s after the first day (later than %s) for a period of %s hours:"
            % (vehicle_type_ID, seconds2date(xmin_calcs), period / 3600)
        )
        labels = ["stress level", "trips", "vehicles"]
        for label, values in zip(labels, [sl, n_trips, n_vehicles]):
            print("\t", label, ":")
            print("\t\tMin: %s" % min(values[xmin_calcs:]))
            print("\t\tMax: %s" % max(values[xmin_calcs:]))
            print("\t\tMean: %s" % np.mean(values[xmin_calcs:]))
            print("\t\tMedian: %s" % np.median(values[xmin_calcs:]))

        if show or save:
            fig, ax = baseplot(show)

            ax.plot(sl)
            ax.plot(n_trips, "--")
            ax.plot(n_vehicles, "--")

            plot_title = (
                "Stress level for vehicle type(s) %s for a period of %s hours"
                % (vehicle_type_ID, period / 3600)
            )
            ax.set_title(plot_title)
            plt.xlim(*self.xlim)
            to_dateaxis(ax)
            plt.grid()
            plt.legend(labels)

            dpi = fig.get_dpi()
            fig.set_size_inches(1920.0 / float(dpi), 948.0 / float(dpi))

            if show:
                fig.show()
            if save:
                basefilename += "_period" + str(period) + "_" + vehicle_type_ID
                basefilename = basefilename.replace(",", "")
                savefig(fig, self.path_results + basefilename, formats)
            if not show:
                plt.close(fig)

    def sl_all(
        self, show=True, save=False, basefilename="stress_level", formats=("pdf",)
    ):
        """Plot the stress level of all vehicle types/ vehicle type
        groups over time as line.
        """
        if not self.gc["depot"]["log_sl"]:
            raise ValueError(
                "stress level (sl) values were not logged. Rerun the simulation "
                "with globalConstants['depot']['log_sl'] = True."
            )
        if not hasattr(self.sl_logs, "calculated"):
            self.calculate_sl()

        self.results["sl"] = {}

        self.results["sl"]["categories"] = globalConstants["depot"][
            "vehicle_type_categories"
        ][::-1]

        sl_values = []
        for category in self.results["sl"]["categories"]:
            sl = discrete2continuous_logs(
                self.depot.evaluation.sl_logs, first=0, subkeys=["sl", category]
            )
            sl_values.append(sl)

        # Set sim time lower bound to avoid inaccuracy of first day
        xmin_calcs = 86400 if self.SIM_TIME > 86400 else 0

        self.results["sl"]["minimums"] = []
        self.results["sl"]["maximums"] = []
        self.results["sl"]["means"] = []
        self.results["sl"]["medians"] = []

        for values in sl_values:
            self.results["sl"]["minimums"].append(min(values[xmin_calcs:]))
            self.results["sl"]["maximums"].append(max(values[xmin_calcs:]))
            self.results["sl"]["means"].append(np.mean(values[xmin_calcs:]))
            self.results["sl"]["medians"].append(np.median(values[xmin_calcs:]))

        # Totals
        self.results["sl"]["minimum"] = min(self.results["sl"]["minimums"])
        self.results["sl"]["maximum"] = max(self.results["sl"]["maximums"])
        self.results["sl"]["mean"] = np.mean(self.results["sl"]["means"])

        if show or save:
            # Print
            period = globalConstants["depot"]["sl_period"]
            print(
                "All current dispatch safety values after the first day (later than %s) for a period of %s hours:"
                % (seconds2date(xmin_calcs), period / 3600)
            )

            for i, category in enumerate(self.results["sl"]["categories"]):
                print("\t%s:" % category)
                print("\t\tMin: %s" % self.results["sl"]["minimums"][i])
                print("\t\tMax: %s" % self.results["sl"]["maximums"][i])
                print("\t\tMean: %s" % self.results["sl"]["means"][i])
                print("\t\tMedian: %s" % self.results["sl"]["medians"][i])

            print("\t Overall:")
            print("\t\tMin: %s" % self.results["sl"]["minimum"])
            print("\t\tMax: %s" % self.results["sl"]["maximum"])
            print("\t\tMean: %s" % self.results["sl"]["mean"])

            # Plot
            fig, ax = baseplot(show)

            for values in sl_values:
                ax.plot(values)

            plot_title = (
                "Current dispatch safety for all vehicle types or vehicle type groups for a period of %s hours"
                % (period / 3600)
            )
            ax.set_title(plot_title)
            plt.xlim(*self.xlim)
            to_dateaxis(ax)
            plt.grid()
            plt.legend(self.results["sl"]["categories"])

            dpi = fig.get_dpi()
            fig.set_size_inches(1920.0 / float(dpi), 948.0 / float(dpi))

            if show:
                fig.show()
            if save:
                basefilename += "_period" + str(period) + "_all"
                basefilename = basefilename.replace(",", "")
                savefig(fig, self.path_results + basefilename, formats)
            if not show:
                plt.close(fig)

    def calc_count_rfd_unblocked_total(self):
        y = np.zeros(self.SIM_TIME, dtype=np.int32)

        for area in self.depot.list_areas:
            if area.issink:
                if "count_rfd_unblocked" not in area.logger.loggedData:
                    print(
                        "Plot 'count_rfd_unblocked' requires logging of "
                        "'count_rfd_unblocked' at areas."
                    )
                    return

            yi = area.logger.get_valList("count_rfd_unblocked", SIM_TIME=self.SIM_TIME)
            yi = np.array(to_prev_values(yi))
            y += yi

        self.results["count_rfd_unblocked_total"] = {
            "min": np.min(y[86400:]),
            "max": np.max(y[86400:]),
            "mean": np.mean(y[86400:]),
        }
        return y

    def count_rfd_unblocked_total(
        self,
        show=True,
        save=False,
        basefilename="count_rfd_unblocked_total",
        formats=("png",),
    ):
        """Plot the total number of ready for departure, unblocked vehicles at
        parking areas over time.
        """
        y = self.calc_count_rfd_unblocked_total()
        if len(y) > 86400:
            print("After the first day:")
            print("\tMin: %d" % self.results["count_rfd_unblocked_total"]["min"])
            print("\tMax: %d" % self.results["count_rfd_unblocked_total"]["max"])
            print("\tMean: %f" % self.results["count_rfd_unblocked_total"]["mean"])

        if show or save:
            fig, ax = baseplot(show)

            ax.plot(y)
            to_dateaxis(ax)

            ax.set_title("Total number of vehicles rfd and unblocked in the depot")
            plt.ylabel("No of vehicles")

            plt.xlim(*self.xlim)

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    # Evaluation and plots specific for areas
    def nvehicles_area(
        self,
        area_ID,
        show=True,
        save=False,
        basefilename="nvehicles_area",
        formats=("png",),
        language="eng",
    ):
        """Plot vehicle count over time on the area with *area_ID*."""
        area = self.depot.areas[area_ID]

        plot_series = area.logger.get_valList("count", SIM_TIME=self.SIM_TIME)
        plot_series = to_prev_values(plot_series)

        if len(plot_series) > 86400:
            print("After the first day:")
            print("\tMin: %d" % np.min(plot_series[86400:]))
            print("\tMax: %d" % np.max(plot_series[86400:]))

        if show or save:
            fig, ax = baseplot(show)
            setting_language(language)

            ax.plot(plot_series)
            to_dateaxis(ax)

            ax.set_title('Number of vehicles at area "%s"' % area.ID)
            if language == "eng":
                plt.ylabel("No of vehicles")
            else:
                plt.ylabel("Anzahl Fahrzeuge")

            plt.xlim(*self.xlim)

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename + "_" + area_ID
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    # Evaluation and plots specific for a group
    def nvehicles_group(
        self,
        group_ID,
        show=True,
        save=False,
        basefilename="nvehicles_group",
        formats=("png",),
        language="eng",
    ):
        """Plot vehicle count over time on the area with *area_ID*."""

        group = self.depot.groups[group_ID]

        if show or save:
            fig, ax = baseplot(show)
            setting_language(language)

            plot_series_total = np.zeros(self.SIM_TIME, dtype=np.int32)

            for area in group.stores:
                plot_series = area.logger.get_valList("count", SIM_TIME=self.SIM_TIME)
                plot_series = to_prev_values(plot_series)
                if len(plot_series) > 86400:
                    print("After the first day:" + str(area))
                    print("\tMin: %d" % np.min(plot_series[86400:]))
                    print("\tMax: %d" % np.max(plot_series[86400:]))

                ax.plot(plot_series, label=area.entry_filter.vehicle_types_str[0])
                plot_series = np.array(plot_series)
                plot_series_total += plot_series

            ax.plot(plot_series_total, label="Total")
            to_dateaxis(ax)
            # ax.set_title('Number of vehicles at area "%s"' % group.ID)
            if language == "eng":
                plt.ylabel("No of vehicles")
            else:
                plt.ylabel("Anzahl Fahrzeuge")

            plt.xlim(*self.xlim)
            plt.legend()

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename + "_" + group_ID
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def count_rfd_unblocked(
        self,
        area_ID,
        show=True,
        save=False,
        basefilename="count_rfd_unblocked",
        formats=("png",),
    ):
        """Plot the number of ready for departure, unblocked vehicles at area
        with *area_ID* over time.
        """
        area = self.depot.areas[area_ID]

        if "count_rfd_unblocked" not in area.logger.loggedData:
            print(
                "Plot 'count_rfd_unblocked' requires logging of "
                "'count_rfd_unblocked' at areas."
            )
            return
        plot_series = area.logger.get_valList(
            "count_rfd_unblocked", SIM_TIME=self.SIM_TIME
        )
        plot_series = to_prev_values(plot_series)

        if len(plot_series) > 86400:
            print("After the first day:")
            print("\tMin: %d" % np.min(plot_series[86400:]))
            print("\tMax: %d" % np.max(plot_series[86400:]))

        if show or save:
            fig, ax = baseplot(show)

            ax.plot(plot_series)
            to_dateaxis(ax)

            ax.set_title('Number of vehicles rfd and unblocked at "%s"' % area.ID)
            plt.ylabel("No of vehicles")

            plt.xlim(*self.xlim)

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename + "_" + area_ID
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    # Evaluation and plots related to the vehicle generator
    def nvehicles_generated(
        self,
        show=True,
        save=False,
        basefilename="nvehicles_generated",
        formats=("png",),
    ):
        """Plot vehicles generated over time."""
        plot_series = self.vehicle_generator.logger.get_valList(
            "count", SIM_TIME=self.SIM_TIME
        )
        plot_series = to_prev_values(plot_series)

        if show or save:
            fig, ax = baseplot(show)

            ax.plot(plot_series)
            ax.set_title("Cumulative number of vehicles generated")
            plt.xlabel("Time [s]")
            plt.ylabel("No of vehicles")
            to_dateaxis(ax)

            adjust_plt()

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    # Evaluation and plots related to the timetable
    def std_line(
        self, show=True, save=False, basefilename="std_line", formats=("png",)
    ):
        """Plot target departure times of trips (without repetition) as line."""
        plot_series = [trip.std for trip in self.timetable.trips]

        if show or save:
            fig, ax = baseplot(show)

            ax.plot(plot_series)
            ax.set_title("Target departure times of trips")
            plt.xlabel("Trip number")
            plt.ylabel("Time [s]")

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def tripattr_scatter(
        self, attr, plot_title, basefilename, show=True, save=False, formats=("png",)
    ):
        """Generic scatter plot for attribute *attr* of class SimpleTrip.
        Without repetition, meaning only trips for which sta is lower than
        timetable.interval_covered are included. *attr* is intended to be a
        time-related integer such as 'std'.
        """
        tdata = [getattr(trip, attr) for trip in self.timetable.trips]
        tdata_clean = [int(t) for t in tdata if t is not None]

        x = list(range(max(tdata_clean) + 1))
        y = [None] * (max(tdata_clean) + 1)
        for trip_no, d_time in enumerate(tdata_clean):
            y[d_time] = trip_no

        if show or save:
            fig, ax = baseplot(show)

            plt.scatter(x, y, marker="x")
            plt.xlabel("Time [s]")
            plt.ylabel("Trip number")
            plt.title(plot_title)
            adjust_plt()

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def std_scatter(
        self, show=True, save=False, basefilename="std_scatter", formats=("png",)
    ):
        """Plot SimpleTrip.std as scatter."""
        self.tripattr_scatter(
            "std",
            "Scheduled departure times of trips",
            basefilename,
            show,
            save,
            formats,
        )

    def sta_scatter(
        self, show=True, save=False, basefilename="sta_scatter", formats=("png",)
    ):
        """Plot SimpleTrip.sta as scatter."""
        self.tripattr_scatter(
            "sta", "Scheduled arrival times of trips", basefilename, show, save, formats
        )

    def ata_scatter(
        self, show=True, save=False, basefilename="ata_scatter", formats=("png",)
    ):
        """Plot SimpleTrip.ata as scatter."""
        self.tripattr_scatter(
            "ata", "Actual arrival times of trips", basefilename, show, save, formats
        )

    def lead_time_match_scatter(
        self,
        show=True,
        save=False,
        basefilename="lead_time_match_scatter",
        formats=("png",),
        language="eng",
    ):
        """Plot SimpleTrip.lead_time_match as scatter.
        Only includes trips that were successfully matched by the dispatch,
        therefore trips that got a vehicle from a depot's init store are not
        included.
        """
        tdata = [trip.lead_time_match for trip in self.timetable.all_trips]
        y = [int(t) / 60 for t in tdata if t is not None]
        x = range(len(y))

        if show or save:
            fig, ax = baseplot(show)
            setting_language(language)

            plt.scatter(x, y, marker="x")

            if language == "eng":
                plt.xlabel("Trip ID")
            else:
                plt.xlabel("Trip ID")

            if language == "eng":
                plt.ylabel("Dispatch time [min]")
            else:
                plt.ylabel("Zeitpunkt Disposition [min]")

            plt.title("Lead times of successful matches in dispatch")

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def lead_time_match_dist(
        self,
        show=True,
        save=False,
        basefilename="lead_time_match_dist",
        formats=("png",),
        language="eng",
    ):
        """Plot the distribution of SimpleTrip.lead_time_match_dist.
        Only includes trips that were successfully matched by the dispatch,
        therefore trips that got a vehicle from a depot's init store are not
        included.
        """
        tdata = [trip.lead_time_match for trip in self.timetable.all_trips]
        y = [int(t) / 60 for t in tdata if t is not None]
        tsm = len(y)  # total number of successful matches
        tsm_lead = y.count(max(y))  # total number of successful matches at lead time

        print(
            "Successful disposition at lead time: " + str(tsm_lead / tsm * 100) + " %"
        )

        if show or save:
            fig, ax = baseplot(show)

            bins = 10
            plt.hist(y, bins=bins)
            plt.xticks(np.arange(0, max(y) + 1, step=(max(y) / bins)))

            if language == "eng":
                plt.xlabel("Time [min]")
            else:
                plt.xlabel("Zeit [min]")

            if language == "eng":
                plt.ylabel("Amount successful disposition")
            else:
                plt.ylabel("Anzahl an Dispositionsvorgängen")

            if language == "eng":
                ax.annotate("n= %d" % tsm, xy=(2, tsm_lead))
            else:
                ax.annotate("n= %d" % tsm, xy=(2, tsm_lead))

            plt.title("Distribution of lead times of successful matches in dispatch")
            plt.xlim(left=0)
            plt.ylim(bottom=0)
            plt.grid()

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def trip_distances_scatter(
        self,
        show=True,
        save=False,
        basefilename="trip_distances_scatter",
        formats=("png",),
    ):
        """Plot trip (without repetition) distance vs departure time."""
        plot_series = [trip.distance for trip in self.timetable.trips]
        t_dep = [trip.std for trip in self.timetable.trips]

        if show or save:
            fig, ax = baseplot(show)

            plt.scatter(t_dep, plot_series, marker="x", linewidths=1)
            plt.xlabel("Departure time [s]")
            plt.ylabel("Trip distance [km]")
            plt.title("Trip distances")
            adjust_plt()

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def trip_distances_dist_bins(
        self,
        bins=15,
        distance_limit=150,
        show=True,
        save=False,
        basefilename="trip_distance_distribution",
        formats=("png",),
    ):
        """Plot trip (without repetition) distance in a histogram.

        bins: [int] number of bins for the histogram; matplotlib parameter.
        distance_limit: [int] in km. The number of trips with higher distance
            is printed.
        """
        plot_series = [trip.distance for trip in self.timetable.trips]

        # Print exact values
        n_crit_trips = 0
        n_trips = len(plot_series)
        for s_tr in plot_series:
            if s_tr >= distance_limit:
                n_crit_trips += 1
        print(
            "No of trips with distance higher than or equal to %s: %s out "
            "of %s, which equals %s %%"
            % (distance_limit, n_crit_trips, n_trips, n_crit_trips / n_trips * 100)
        )

        if show or save:
            fig, ax = baseplot(show)

            plt.hist(plot_series, bins=bins)
            plt.xlabel("Trip distance [km]")
            plt.ylabel("Number of trips")
            plt.title("Trip distance distribution")
            adjust_plt_hist()

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def trip_distances_dist(
        self,
        show=True,
        save=False,
        basefilename="trip_distance_distribution",
        formats=("png",),
        language="eng",
    ):
        """Plot trip (without repetition) distance in a histogram."""
        vehicle_types = list(self.gc["depot"]["vehicle_types"].keys())

        for vt in vehicle_types:
            distances = [
                trip.distance
                for trip in self.timetable.trips
                if trip.vehicle_types_joinedstr == vt
            ]
            distances.sort()

            if show or save:
                fig, ax = baseplot(show, figsize=(10, 8))
                setting_language(language)

                if language == "eng":
                    plt.ylabel("Distance [km]")
                else:
                    plt.ylabel("Umlauflänge [km]")

                if language == "eng":
                    plt.xlabel("Trip")
                else:
                    plt.xlabel("Umlauf")

                plt.xlim(left=0, right=len(distances))
                plt.ylim(bottom=0, top=max(distances))

                plt.bar(np.arange(len(distances)), distances, width=0.5)

                if show:
                    fig.show()
                if save:
                    filename = self.path_results + basefilename
                    savefig(fig, filename, formats)
                if not show:
                    plt.close(fig)

    def target_vehicle_types_dist(
        self,
        show=True,
        save=False,
        basefilename="target_vehicle_types_distribution",
        formats=("png",),
    ):
        """Plot the distribution of vehicle_types of trips (without
        repetition) in a histogram.
        """
        occurrences = [trip.vehicle_types_joinedstr for trip in self.timetable.trips]
        vehicle_types = list(set(occurrences))
        typemap = {}
        for i, vehicletype in enumerate(vehicle_types):
            typemap[vehicletype] = i

        plot_series = [typemap[vType] for vType in occurrences]

        counter = Counter(occurrences)
        print("Occurences: %s" % dict(counter))

        if show or save:
            fig, ax = baseplot(show)

            plt.hist(plot_series, bins=len(vehicle_types))
            plt.xlabel("Vehicle Type")
            plt.ylabel("Number of trips")
            plt.title("Target vehicle type distribution")

            x_steps = range(len(vehicle_types))
            plt.xticks(x_steps, vehicle_types)

            adjust_plt_hist()

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def departure_delay_vt_calculation(self, vehicle_types=None):
        if vehicle_types is None:
            vehicle_types = list(self.gc["depot"]["vehicle_types"].keys())

        now = self.timetable.env.now
        delaydata = {}
        for vt in vehicle_types:
            delaydata[vt] = {"all": [], "total_trip_count": 0}
        n_trips_finished = 0

        for trip in self.timetable.trips_issued:
            if trip.ata is not None:
                n_trips_finished += 1
                for vt in vehicle_types:
                    if trip.vehicle.vehicle_type.ID == vt:
                        delaydata[vt]["total_trip_count"] += 1
                        if trip.atd is not None:
                            # Trip has started
                            delay = trip.atd - trip.std
                            delaydata[vt]["all"].append(delay)
                        elif trip.std < now:
                            # Trip should have started before simulation end
                            delay = now - trip.std
                            delaydata[vt]["all"].append(delay)
                        else:
                            # Trip is not due yet. Append None to keep the data
                            # plottable by trip
                            delaydata[vt]["all"].append(None)
                    else:
                        # Append None for other vehicle types to keep the data
                        # plottable by trip
                        delaydata[vt]["all"].append(None)
        return delaydata, n_trips_finished

    def departure_delay_vt(
        self,
        vehicle_types=None,
        show=True,
        save=False,
        basefilename="departure_delay_vt",
        formats=("png",),
    ):
        """Delay of trips (with repetition) upon sim time end for selected
        target vehicle types, for all depots. Plot as scatter. Includes
        finished trips only!

        vehicle_types: [list] of vehicle type IDs. If None, evaluate for all
            types.
        """
        delaydata, n_trips_finished = self.departure_delay_vt_calculation(
            vehicle_types=vehicle_types
        )

        # Evaluation that will be printed in the console
        for vt in delaydata:
            values_only = [t for t in delaydata[vt]["all"] if t is not None]

            delaydata[vt]["sum"] = 0
            delaydata[vt]["count"] = 0
            delaydata[vt]["mean"] = 0
            delaydata[vt]["max"] = max(values_only) if values_only else 0
            for entry in delaydata[vt]["all"]:
                if entry is not None and entry != 0:
                    delaydata[vt]["sum"] += entry
                    delaydata[vt]["count"] += 1

            if delaydata[vt]["count"] > 1:
                delaydata[vt]["mean"] = delaydata[vt]["sum"] / delaydata[vt]["count"]
            if delaydata[vt]["total_trip_count"] > 0:
                # Share of delayed trips
                delaydata[vt]["share_delayed"] = (
                    delaydata[vt]["count"] / delaydata[vt]["total_trip_count"] * 100
                )
                # Mean of delays including not delayed
                delaydata[vt]["mean_all"] = (
                    delaydata[vt]["sum"] / delaydata[vt]["total_trip_count"]
                )
            else:
                delaydata[vt]["share_delayed"] = 0
                delaydata[vt]["mean_all"] = 0

        sum_all = 0
        count_all = 0
        count_delayed_all = 0
        max_all = 0
        for vt in delaydata:
            sum_all += delaydata[vt]["sum"]
            count_all += delaydata[vt]["total_trip_count"]
            count_delayed_all += delaydata[vt]["count"]
            max_all = max(max_all, delaydata[vt]["max"])
        share_delayed_all = count_delayed_all / count_all * 100 if count_all > 0 else 0
        average_all = sum_all / count_all
        average_all_delayed = (
            sum_all / count_delayed_all if count_delayed_all > 0 else 0
        )

        for vt in delaydata:
            print("\nVehicle type %s:" % vt)
            print(
                "\ttotal number of finished trips: %d"
                % delaydata[vt]["total_trip_count"]
            )
            print(
                "\tnumber delayed trips: %d (%.2f %%)"
                % (delaydata[vt]["count"], round(delaydata[vt]["share_delayed"], 2))
            )
            print("\tmax delay: %d min" % (delaydata[vt]["max"] / 60))
            print("\tdelay sum: %d min" % (delaydata[vt]["sum"] / 60))
            print(
                "\tmean delay including trips on time: %d min"
                % (delaydata[vt]["mean_all"] / 60)
            )
            print(
                "\tmean delay only for delayed trips: %d min"
                % (delaydata[vt]["mean"] / 60)
            )

        print("\nOverall:")
        print("\ttotal number of finished trips: %d" % count_all)
        print(
            "\tnumber of delayed trips: %d (%.2f %%)"
            % (count_delayed_all, share_delayed_all)
        )
        print("\tmax delay: %d min" % (max_all / 60))
        print("\tdelay sum: %d min" % (sum_all / 60))
        print("\tmean delay including trips on time: %d min" % (average_all / 60))
        print(
            "\tmean delay only for delayed trips: %d min" % (average_all_delayed / 60)
        )

        if show or save:
            fig, ax = baseplot(show)

            x = list(range(n_trips_finished))
            for vt in delaydata:
                # Replacement for not showing non-delayed
                plot_series = [None if t == 0 else t for t in delaydata[vt]["all"]]
                plot_series_min = [None if i is None else i / 60 for i in plot_series]
                plt.scatter(x, plot_series_min, marker="x", label=vt)

            plt.xlabel("Trip number")
            plt.ylabel("Delay [min]")
            plt.title("Delay of trips upon departure by target vehicle type")
            plt.legend()

            # adjust_plt()
            plt.grid(True)

            plt.ylim(0, max_all / 60 + 10)

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def total_delay(self):
        """Save the total delay sum of all trips that were due until sim time
        end.
        """
        self.results["total_delay"] = 0
        for trip in self.timetable.trips_issued:
            if trip.std < self.timetable.env.now and (
                trip.atd is None or trip.atd > trip.std
            ):
                self.results["total_delay"] += trip.delay_departure

    # Evaluation and plots specific for vehicles
    def area_history(
        self,
        vehicle_ID,
        show=True,
        save=False,
        basefilename="area_history",
        formats=("png",),
    ):
        """Plot a vehicle's location history on depot areas similar to a gantt
        chart. Includes used processes only.
        """
        vehicle = self.vehicle_generator.select(vehicle_ID)
        if vehicle is None:
            raise ValueError("Vehicle with ID '%s' couldn't be found" % vehicle_ID)

        plot_title = "Vehicle %s (type %s) area location history" % (
            vehicle_ID,
            vehicle.vehicle_type.ID,
        )

        data_dict = vehicle.logger.loggedData["dwd.current_area"]

        plotdata = {}
        previous = None

        # Determine start and end times
        for t in data_dict:
            if data_dict[t]:
                current = data_dict[t]
                if current == previous:
                    continue

                if current.ID not in plotdata:
                    plotdata[current.ID] = {}
                    plotdata[current.ID]["ref"] = current
                    plotdata[current.ID]["starts"] = []
                    plotdata[current.ID]["ends"] = []
                    plotdata[current.ID]["xranges"] = []

                plotdata[current.ID]["starts"].append(t)

                if previous:
                    plotdata[previous.ID]["ends"].append(t)

                previous = current
            else:
                if previous:
                    plotdata[previous.ID]["ends"].append(t)
                previous = None

        # Fill xranges with tuples (start, duration)
        for ID in plotdata:
            for startno, start in enumerate(plotdata[ID]["starts"]):
                if startno != len(plotdata[ID]["ends"]):
                    # startno has a match in ends
                    plotdata[ID]["xranges"].append(
                        (start, plotdata[ID]["ends"][startno] - start)
                    )
                else:
                    # ends is shorter than starts; proc lasted until simtime
                    plotdata[ID]["xranges"].append((start, self.SIM_TIME - start))

        # # Sort plotdata by first start value
        # plotdata = OrderedDict(sorted(plotdata.items(),
        #                               key=lambda v: v[1]['xranges'][0][0]))

        # Sort plotdata by area appearance in plan and set yrange
        areas_chron = {}
        pos = 0

        def check_entry(entry, sorted_areas, pos):
            """Summary to gather entries and assign position"""
            if entry.ID not in sorted_areas:
                pos += 1
                sorted_areas[entry.ID] = pos
            return sorted_areas, pos

        for entry in self.depot.default_plan:
            if isinstance(entry, eflips.depot.BaseArea):
                areas_chron, pos = check_entry(entry, areas_chron, pos)
            elif isinstance(entry, eflips.depot.AreaGroup):
                for subentry in entry.stores:
                    areas_chron, pos = check_entry(subentry, areas_chron, pos)

        plotdata = OrderedDict(
            sorted(plotdata.items(), key=lambda pair: areas_chron[pair[0]])
        )

        ID_count = 0
        for ID in plotdata:
            ID_count += 1
            plotdata[ID]["IDno"] = ID_count
            plotdata[ID]["yrange"] = (ID_count * 10, 9.8)

        # Plot
        if show or save:
            fig, ax = baseplot(show)

            for ID in plotdata:
                print(ID + ":")
                print("\txranges: ", plotdata[ID]["xranges"])
                print("\tyrange: ", plotdata[ID]["yrange"])
                ax.broken_barh(plotdata[ID]["xranges"], plotdata[ID]["yrange"])

            ax.set_ylim(5, len(plotdata) * 10 + 15)
            ax.set_xlim(*self.xlim)
            ax.set_xlabel("Time [s]")
            ax.set_yticks([plotdata[ID]["IDno"] * 10 + 5 for ID in plotdata])
            ax.set_yticklabels(plotdata.keys())
            ax.grid(True)
            plt.title(plot_title)

            if show:
                fig.show()
            if save:
                filename = (
                    self.path_results + "vehicle_" + vehicle_ID + "_" + basefilename
                )
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def vehicle_type_dist(
        self,
        show=True,
        save=False,
        basefilename="vehicle_type_distribution",
        formats=("png",),
    ):
        """Plot the distribution of vehicle types in a histogram."""
        occurrences = [
            vehicle.vehicle_type.ID for vehicle in self.vehicle_generator.items
        ]
        vehicle_types = list(set(occurrences))
        typemap = {}
        for vehicletypeNo, vehicletype in enumerate(vehicle_types):
            typemap[vehicletype] = vehicletypeNo

        plot_series = [typemap[vType] for vType in occurrences]
        print("typemap: ", typemap)
        print("plot_series: ", plot_series)
        np.arange(len(vehicle_types))
        np.arange(len(vehicle_types)) + 0.5

        # Also print exact values
        count = Counter(occurrences)
        total = sum(count.values())
        print("Count of vehicle types: %s" % count)
        print("Total: %d" % total)

        if show or save:
            fig, ax = baseplot(show)

            plt.hist(plot_series, bins=len(vehicle_types))
            # plt.hist(plot_series, bins=np.arange(len(vehicle_types)) + 0.5)
            plt.xlabel("Vehicle Type")
            plt.ylabel("Number of vehicles")
            plt.title("Vehicle type distribution")

            x_steps = range(len(vehicle_types))
            plt.xticks(x_steps, vehicle_types)
            adjust_plt_hist()

            # x_steps = range(len(vehicle_types) + 1)
            # plt.xticks(x_steps, vehicle_types + [''])
            # plt.xlim([-1, len(vehicle_types)])

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def processhistory_single_vehicle_separate(
        self,
        vehicle_ID,
        show=True,
        save=False,
        basefilename="processhistory_single_vehicle_separate",
        formats=("png",),
    ):
        """Plot the usage history of processes by a vehicle in a horizontal
        broken bar chart. Includes only used processes on the y axis. Process
        times exclude potential waiting for resources.
        """
        vehicle = self.vehicle_generator.select(vehicle_ID)
        if vehicle is None:
            raise ValueError("Vehicle with ID '%s' couldn't be found" % vehicle_ID)

        data_dict = vehicle.logger.loggedData["dwd.active_processes_copy"]

        plotdata = {}
        used_ID_count = 0
        registered_instances = []

        for t in data_dict:
            for proc in data_dict[t]:
                if proc not in registered_instances:
                    if proc.ID not in plotdata:
                        used_ID_count += 1
                        plotdata[proc.ID] = {}
                        plotdata[proc.ID]["IDno"] = used_ID_count
                        plotdata[proc.ID]["xranges"] = []
                        plotdata[proc.ID]["yrange"] = (used_ID_count * 10, 9.8)

                    # Fill xranges with tuples (start, duration)
                    for startno, start in enumerate(proc.starts):
                        if startno != len(proc.ends):
                            # startno has a match in end
                            plotdata[proc.ID]["xranges"].append(
                                (start, proc.ends[startno] - start)
                            )
                        else:
                            # ends is shorter than starts; proc lasted until simtime
                            plotdata[proc.ID]["xranges"].append(
                                (start, self.SIM_TIME - start)
                            )

        # Sort plotdata by first start value
        plotdata = OrderedDict(
            sorted(plotdata.items(), key=lambda v: v[1]["xranges"][0][0])
        )

        if show or save:
            fig, ax = baseplot(show)

            for ID in plotdata:
                print(ID + ":")
                print("\txranges: ", plotdata[ID]["xranges"])
                print("\tyrange: ", plotdata[ID]["yrange"])
                ax.broken_barh(plotdata[ID]["xranges"], plotdata[ID]["yrange"])
                # facecolors = 'blue'

            ax.set_ylim(5, len(plotdata) * 10 + 15)
            ax.set_xlim(*self.xlim)
            ax.set_xlabel("Time [s]")
            ax.set_yticks([plotdata[ID]["IDno"] * 10 + 5 for ID in plotdata])
            ax.set_yticklabels(plotdata.keys())
            ax.grid(True)
            to_dateaxis(ax)
            plt.title(
                'Vehicle "%s" (type %s) process history'
                % (vehicle_ID, vehicle.vehicle_type.ID)
            )

            if show:
                fig.show()
            if save:
                filename = (
                    self.path_results + "vehicle_" + vehicle_ID + "_" + basefilename
                )
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def process_intervals(
        self,
        process_ID,
        crit_value=86400 * 2,
        bins=15,
        show=True,
        save=False,
        basefilename="process_intervals",
        formats=("png",),
    ):
        """Distribution of time intervals between the application of process
        with ID *process_ID* on vehicles.

        bins: [int] number of bins for the histogram; matplotlib parameter.
        The interval from 0 to the first appearance is skipped.
        Prints appearances if > *crit_value*.
        """
        # Check if process exists
        if process_ID not in self.depot.processes:
            raise ValueError(
                "Process with ID '%s' not found in depot '%s'"
                % (process_ID, self.depot.ID)
            )

        # Calculate for each vehicle
        intervals_by_v = {v.ID: {"intervals": []} for v in self.vehicle_generator.items}

        for v in self.vehicle_generator.items:
            prev = None
            for processes in v.logger.loggedData["dwd.active_processes_copy"].values():
                for proc in processes:
                    if proc.ID == process_ID:
                        if prev is not None:  # first appearance is skipped
                            intervals_by_v[v.ID]["intervals"].append(
                                proc.ends[-1] - prev.ends[-1]
                            )
                        prev = proc

        # Extra calc
        sum_crits = 0
        v_with_crit = []

        for vID, d in intervals_by_v.items():
            calc_descr_stats(d, "intervals")

            # count values > 2 days
            d["above_crit_value"] = sum(i > crit_value for i in d["intervals"])
            if d["above_crit_value"]:
                sum_crits += d["above_crit_value"]
                v_with_crit.append(vID)

        # Summarize
        intervals_all = {"intervals": []}
        for i in intervals_by_v.values():
            intervals_all["intervals"].extend(i["intervals"])
        calc_descr_stats(intervals_all, "intervals")

        # Output for each vehicle
        print("\n Intervals of process {} for each vehicle:".format(process_ID))
        # for vID, values in intervals_by_v.items():
        #     print('Vehicle {:8s}: '.format(vID),
        #           'min: {:7d}'.format(values['min']),
        #           ', max: ', values['max'],
        #           ', mean: %.0f' % round(values['mean']),
        #           ', median: %.0f' % round(values['median']),
        #           'intervals: ', values['intervals']
        #           )
        pp.pprint(intervals_by_v)

        # Output summary
        print(
            "\n Summary of all intervals: ",
            "min: ",
            intervals_all["min"],
            ", max: ",
            intervals_all["max"],
            ", mean: %.0f" % round(intervals_all["mean"]),
            ", median: %.0f" % round(intervals_all["median"]),
        )

        print(
            "\n Number of intervals > %s: %d (= %.2f %%)"
            % (crit_value, sum_crits, sum_crits / len(intervals_all))
        )
        print("vehicles: %s" % v_with_crit)

        if show or save:
            fig, ax = baseplot(show)

            plt.hist(intervals_all["intervals"], bins=bins)
            plt.xlabel("time [s]")
            plt.title(
                "Distribution of intervals of process '%s' for all "
                "vehicles" % process_ID
            )
            # plt.xlim(left=0)
            plt.ylim(bottom=0)
            plt.grid()

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def battery_level(
        self,
        vehicle_ID,
        show=True,
        save=False,
        basefilename="battery_level",
        formats=("png",),
    ):
        """Plot a vehicle's battery level over time as a line.
        Assume that all charging and discharging activities are linear. Might
        be inaccurate if multiple charging/discharging processes were active at
        the same time on the same vehicle.
        """
        vehicle = self.vehicle_generator.select(vehicle_ID)
        if vehicle is None:
            raise ValueError("Vehicle with ID '%s' couldn't be found" % vehicle_ID)

        y = np.zeros(self.SIM_TIME)

        previous = None
        for current in vehicle.battery_logs:
            if previous is None:  # True for first log
                previous = current
                continue

            y[previous.t : current.t] = np.linspace(
                previous.energy, current.energy, current.t - previous.t, endpoint=False
            )
            previous = current

        y[0 : vehicle.battery_logs[0].t] = vehicle.battery_logs[0].energy
        y[vehicle.battery_logs[-1].t :] = vehicle.battery_logs[-1].energy

        if show or save:
            fig, ax = baseplot(show)

            ax.plot(y)
            plt.xlim(*self.xlim)
            # plt.ylim(0, vehicle.battery.energy_real)
            plt.grid(True)
            plt.title(
                "Vehicle %s (type %s) battery level history"
                % (vehicle_ID, vehicle.vehicle_type.ID)
            )
            to_dateaxis(ax)

            if show:
                fig.show()
            if save:
                filename = (
                    self.path_results + "vehicle_" + vehicle_ID + "_" + basefilename
                )
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def vehicle_power(
        self,
        vehicle_ID,
        show=True,
        save=False,
        basefilename="vehicle_power",
        formats=("png",),
    ):
        """Plot the power for a vehicle at charging interfaces during charging
        and preconditioning over time as line.
        """
        vehicle = self.vehicle_generator.select(vehicle_ID)
        if vehicle is None:
            raise ValueError("Vehicle with ID '%s' couldn't be found" % vehicle_ID)

        plot_title = "Vehicle %s power at charging interfaces" % vehicle.ID

        print(vehicle, vehicle.power_logs)
        y = discrete2continuous_logs(vehicle.power_logs, first=0)
        print("Min power: ", min(y))
        print("Max power: ", max(y))

        if show or save:
            fig, ax = baseplot(show)

            ax.plot(y)
            plt.title(plot_title)
            plt.ylabel("Power [kW]")
            plt.xlim(*self.xlim)
            to_dateaxis(ax)
            plt.grid()

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    # Evaluation and plots specific for resources
    def usage_history(
        self,
        resource_ID,
        show=True,
        save=False,
        basefilename="usage_history",
        formats=("png",),
    ):
        """Plot the usage history of a resource.."""
        resource = self.depot.resources[resource_ID]

        user_counts = {}
        for t in resource.logger.loggedData["user_count"]:
            users = resource.logger.loggedData["user_count"][t].keys()
            if users:
                for user in users:
                    if user not in user_counts:
                        # First entry for this user
                        user_counts[user] = [None] * self.SIM_TIME
                    # Get count of this user
                    user_counts[user][t] = resource.logger.loggedData["user_count"][t][
                        user
                    ]
                # Explicitly set count of unmentioned users to zero
                for counted_user in user_counts:
                    if counted_user not in users:
                        user_counts[counted_user][t] = 0
            else:
                # Explicitly set count to zero for all registered users
                for counted_user in user_counts:
                    user_counts[counted_user][t] = 0

        if show or save:
            fig, ax = baseplot(show)

            for user in user_counts:
                # plot.addSeries(, label=user)
                ax.plot(to_prev_values(user_counts[user]), label=user)

            plt.xlabel("Time [s]")
            plt.ylabel("Area")
            plt.title("Resource %s usage history" % resource.ID)
            plt.legend(loc="best")
            plt.ylabel("Count")
            # setting_language('de')
            to_dateaxis(ax)

            adjust_plt()

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def congestion_calulation(self):
        """Calculates the congestion in the depot.
        return: WATCH out times are in hours
        """
        congestion_vehicle = pd.Series()
        areas = ["Pre Depot"]
        for area, value in self.depot.areas.items():
            areas.append(value.ID)
        congestion_area_caused = pd.Series(0.0, index=areas)
        congestion_area_stands = pd.Series(0.0, index=areas)
        total_congestion = 0.0
        total_dwelltime = 0.0

        for vehicle in self.vehicle_generator.items:
            congestion_vehicle[vehicle.ID] = 0.0  # init
            pre_area = "Pre Depot"
            for key, item in vehicle.logger.loggedData["area_waiting_time"].items():
                inter_interval = intersection(
                    self.xlim, [key - item["waiting_time"], key]
                )
                if isinstance(inter_interval, list):
                    waiting_time = inter_interval[1] - inter_interval[0]
                    if waiting_time != 0:
                        waiting_time = waiting_time / 60 / 60  # in h
                        congestion_vehicle[item["vehicle"]] += waiting_time
                        congestion_area_caused[item["area"]] += waiting_time
                        congestion_area_stands[pre_area] += waiting_time
                        total_congestion += waiting_time
                pre_area = item["area"]

            pre_ata = None
            for trip in vehicle.finished_trips:  # Time in depot
                if pre_ata is not None:
                    inter_interval = intersection(self.xlim, [pre_ata, trip.atd])
                    if isinstance(inter_interval, list):
                        total_dwelltime += (
                            (inter_interval[1] - inter_interval[0]) / 60 / 60
                        )
                pre_ata = trip.ata

        return (
            total_congestion,
            congestion_area_caused,
            congestion_area_stands,
            congestion_vehicle,
            total_dwelltime,
        )

    def congestion(
        self,
        show=True,
        save=False,
        basefilename_vehicle="congestion_by_vehicle",
        basefilename_vehicle_sort="congestion_by_vehicle_sort",
        basefilename_area_caused="congestion_by_area_caused",
        basefilename_area_stands="congestion_by_area_stands",
        formats=("png",),
        language=("eng"),
    ):
        """
        Calculates the congestion in the depot. Congestion is the time between an vehicle wants to go to an area
        and when does it arrives
        :param show:
        :param save:
        :param basefilename_vehicle:
        :param basefilename_area_caused:
        :param basefilename_area_stands:
        :param formats:
        :return:
        """

        (
            total_congestion,
            congestion_area_caused,
            congestion_area_stands,
            congestion_vehicle,
            total_dwelltime,
        ) = self.congestion_calulation()

        print("\nTotal congestion in time intervall: %.2f h" % total_congestion)
        print(
            "Percentage of total congestion time on total dwelltime in time intervall %.2f %%"
            % ((total_congestion / total_dwelltime) * 100)
        )
        print("mean: %.2f min" % (congestion_vehicle.mean() * 60))

        if show or save:
            # Plot: Congestion by vehicle
            fig, ax = baseplot(show)
            setting_language(language)

            if language == "eng":
                plt.xlabel("Vehicle ID")
            else:
                plt.xlabel("Fahrzeug ID")

            if language == "eng":
                plt.ylabel("Time")
            else:
                plt.ylabel("Zeit [h]")

            plt.title("Congestion by vehicle")
            ax = plt.gca()
            congestion_vehicle.plot(kind="bar", ax=ax, fontsize=3)

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename_vehicle
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

            # Plot: Congestion by vehicle (sort)
            fig, ax = baseplot(show)

            if language == "eng":
                plt.xlabel("Amount Vehicles")
            else:
                plt.xlabel("Anzahl Fahrzeuge")

            if language == "eng":
                plt.ylabel("Time [h]")
            else:
                plt.ylabel("Zeit [h]")

            plt.title("Congestion by vehicle (sort)")
            ax = plt.gca()
            congestion_vehicle.sort_values(ascending=False, inplace=True)
            # new_index for x-Labels 10 ,20, 30, ..
            new_index = [
                i + 1 if (i + 1) % 10 == 0 else ""
                for i in range(len(congestion_vehicle))
            ]
            congestion_vehicle.index = new_index
            congestion_vehicle.plot(kind="bar", ax=ax, fontsize=10)

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename_vehicle_sort
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

            # Plot: Congestion by area
            fig, ax = baseplot(show)
            if language == "eng":
                plt.xlabel("Area")
            else:
                plt.xlabel("Area")

            if language == "eng":
                plt.ylabel("Time")
            else:
                plt.ylabel("Zeit [h]")

            plt.title("Congestion by area - which area caused the congestion?")
            ax = plt.gca()
            congestion_area_caused.plot(kind="bar", ax=ax)

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename_area_caused
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

            fig, ax = baseplot(show)
            if language == "eng":
                plt.xlabel("Area")
            else:
                plt.xlabel("Area")

            if language == "eng":
                plt.ylabel("Time")
            else:
                plt.ylabel("Zeit [h]")

            plt.title("Congestion by area - on which area does the vehicles stands?")
            ax = plt.gca()
            congestion_area_stands.plot(kind="bar", ax=ax)

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename_area_stands
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def total_parking_congestion(self):
        """Calculate and save the total time of successful and pending put
        requests to parking area groups.
        """
        total = 0

        for pa_group in self.depot.parking_area_groups:
            # Successful requests
            for store in pa_group.stores:
                total += sum(store.tus_put)

            # Requests still pending at sim time end
            for pending_req in pa_group.put_queue:
                total += self.env.now - pa_group.put_queue[pending_req]

        self.results["total_parking_congestion"] = total

    def occupancy_rate_calculation(self):
        """
        Calculates the occupancy rate for each slot.
        :return: dict with occupancy rates in seconds
        """

        areas_and_slots = {}
        for area in self.depot.areas.values():
            areas_and_slots[area.ID] = pd.Series(0, index=range(1, area.capacity + 1))

        for vehicle in self.vehicle_generator.items:
            pre_log = {"area": None, "slot": None, "time": None}
            for time, slot in vehicle.logger.loggedData["dwd.current_slot"].items():
                if pre_log["slot"] is not None and pre_log["area"] is not None:
                    inter_interval = intersection(self.xlim, [pre_log["time"], time])
                    if isinstance(inter_interval, list):
                        time_delta = inter_interval[1] - inter_interval[0]
                        areas_and_slots[pre_log["area"]].loc[
                            pre_log["slot"]
                        ] += time_delta

                if isinstance(
                    vehicle.logger.loggedData["dwd.current_area"][time],
                    eflips.depot.depot.BaseArea,
                ):
                    pre_log["area"] = vehicle.logger.loggedData["dwd.current_area"][
                        time
                    ].ID
                else:
                    pre_log["area"] = vehicle.logger.loggedData["dwd.current_area"][
                        time
                    ]
                pre_log["slot"] = slot
                pre_log["time"] = time

        return areas_and_slots

    def occupancy_rate(self, basefilename="occupancy_rate.xlsx"):
        """
        Writes everything in an excel.
        """
        areas_and_slots = self.occupancy_rate_calculation()

        filename = self.path_results + basefilename
        wb = xlsxwriter.Workbook(filename)
        absolute_sheet = wb.add_worksheet("absolute occupancy")
        percentage_sheet = wb.add_worksheet("percentage occupancy")

        metadata = [
            "time",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "templatename",
            self.configurator.templatename,
            "depot ID",
            self.depot.ID,
        ]
        absolute_sheet.write_row(0, 0, metadata)
        percentage_sheet.write_row(0, 0, metadata)

        max_slot = 0
        i = 1
        per_depot = 0
        absolute_depot = 0
        per_format = wb.add_format()
        per_format.set_num_format(10)  # percentage format
        for key, series in areas_and_slots.items():
            column = [key, series.sum()] + list(series)
            absolute_sheet.write_column(row=4, col=i, data=column)
            absolute_depot += series.sum()

            per_series = series / (self.xlim[1] - self.xlim[0])
            per_area = per_series.sum() / series.__len__()
            column = [key, per_area] + list(per_series)
            percentage_sheet.write_column(
                row=4, col=i, data=column, cell_format=per_format
            )
            per_depot += per_area  # dont forget to divide by number of areas

            i += 1
            max_slot = max(max_slot, series.__len__())

        index_per = ["average occupation (area)"]
        index_abs = ["cumulated sum (area)"]
        for i in range(1, max_slot):
            index_abs.append("slot_" + str(i))
            index_per.append("slot_" + str(i))
            absolute_sheet.write_column(row=5, col=0, data=index_abs)
            percentage_sheet.write_column(row=5, col=0, data=index_per)

        absolute_sheet.write_row(
            row=2, col=0, data=["cumulated sum (all slots in depot)", absolute_depot]
        )
        percentage_sheet.write_row(
            row=2,
            col=0,
            data=[
                "average occupation (all slots in depot)",
                per_depot / areas_and_slots.__len__(),
            ],
            cell_format=per_format,
        )

        wb.close()

    def blocked_slots(
        self, show=True, save=False, basefilename="blocked_slots", formats=("png",)
    ):
        """
        Calculates the blocked slots on LineAreas
        :param show:
        :param save:
        :param basefilename:
        :param formats:
        :return:
        """

        count_parked_vehicle = np.zeros(self.SIM_TIME, dtype=np.int32)
        count_blocked_slots = np.zeros(self.SIM_TIME, dtype=np.int32)
        for area in self.depot.areas.values():
            if isinstance(area, eflips.depot.LineArea):
                nv = area.logger.get_valList("count", SIM_TIME=self.SIM_TIME)
                nv = to_prev_values(nv)
                nv = np.array(nv)
                count_parked_vehicle += nv

                nv = area.logger.get_valList("vacant_blocked", SIM_TIME=self.SIM_TIME)
                nv = to_prev_values(nv)
                nv = np.array(nv)
                count_blocked_slots += nv

        count_together = count_blocked_slots + count_parked_vehicle

        print(
            "Max. number of used and blocked slots (LineArea): "
            + str(count_together.max())
        )
        print(
            "Max. number of blocked slots(LineArea): " + str(count_blocked_slots.max())
        )

        if show or save:
            fig, ax = baseplot(show)

            plt.xlabel("Time [h]")
            plt.ylabel("Total number of blocked slots", color="black")

            plt.title("Blocked slots on Line Areas")

            # Convert x axis seconds to dates
            to_dateaxis(ax)
            plt.xlim(*self.xlim)

            ax.plot(count_blocked_slots, color="black")

            ax2 = ax.twinx()
            ax2.plot(count_together, color="green")
            plt.ylabel("Total number of used and blocked slots", color="green")

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def number_of_resources_calculation(self):
        """
        return: dict with number of charging interfaces and the sum of the capacity of other resources
        """
        counter = {"ci": 0, "capacity_other_resources": 0}

        for resource in self.depot.resources.values():
            if isinstance(resource, eflips.depot.DepotChargingInterface):
                counter["ci"] += 1
            else:
                counter["capacity_other_resources"] += resource.capacity

        return counter

    def pssmart2_ratings(
        self, show=True, save=False, basefilename="pssmart2_ratings", formats=("png",)
    ):
        """Plot logged ratings of PSSmart2 put requests as scatter and print
        some results.
        """
        x = []
        y = []
        x_none = []
        y_none = []
        for group in self.depot.parking_area_groups:
            for time, values in group.pssmart2_logs.items():
                for value in values:
                    if value is not None:
                        x.append(time)
                        y.append(value)
                    else:
                        x_none.append(time)
                        y_none.append(0)
        if not x and not y:
            raise ValueError(
                "park_ratings requires data in ParkingAreaGroup.pssmart2_logs."
            )

        self.results["pssmart2_ratings"] = {}
        self.results["pssmart2_ratings"]["avg"] = np.mean(y)
        self.results["pssmart2_ratings"]["invalids"] = bool(y_none)

        if show or save:
            print("PSSmart2 rating valid values")
            print("\tmin: %s" % min(y))
            print("\tmax: %s" % max(y))
            print("\tmean: %s" % np.mean(y))
            print("\tmedian: %s" % np.median(y))
            print(
                "Number of invalid values (no alternatives at the time of rating): %d"
                % len(x_none)
            )

            fig, ax = baseplot(show, figsize=(16, 9))

            valid = plt.scatter(x, y)
            invalid = plt.scatter(x_none, y_none, marker="x", color="red")

            plt.legend((valid, invalid), ("Valid", "Invalid"))

            ax.set_title("PSSmart2 ratings")
            plt.xlim(*self.xlim)
            plt.ylabel("Value")
            to_dateaxis(ax)

            if show:
                fig.show()
            if save:
                filename = self.path_results + basefilename
                savefig(fig, filename, formats)
            if not show:
                plt.close(fig)

    def validate(self):
        """Calculation of the energy required in the depot in various ways. Work in progress"""

        list_power_logs = list(self.power_logs.items())
        energy = 0
        for i in range(len(list_power_logs)):
            if i != 0:
                energy += (
                    list_power_logs[i][0] - list_power_logs[i - 1][0]
                ) * list_power_logs[i - 1][1]
        print(energy / 60 / 60)


def baseplot(show, figsize=None):
    """Return new fig and ax after setting interactive mode based on *show*
    [bool].
    figsize: [None or tuple] figure size in cm (not inches)
    """
    if show:
        plt.ion()
        plt.show()  # may be required to return to the console
    else:
        plt.ioff()

    if figsize is not None:
        figsize = cm2in(*figsize)

    fig, ax = plt.subplots(figsize=figsize)
    return fig, ax


def setting_language(language):
    # English (default)
    if language == "eng":
        locale.setlocale(locale.LC_ALL, "en_US")

    # German
    if language == "de":
        locale.setlocale(locale.LC_ALL, "de_DE")


def intersection(interval_a, interval_b):
    """

    :param interval_a: list of two numbers
    :param interval_b: list of two numbers
    :return: the intersection of the two intervalls or False if no intersection is there
    """
    lower = max(interval_a[0], interval_b[0])
    upper = min(interval_a[1], interval_b[1])
    if lower >= upper:
        return False
    else:
        return [lower, upper]


def savefig(fig, filename, formats=("png",), confirm=True, dpi=None, **kwargs):
    """filename: [str] including path, excluding extension.

    formats: [tuple] of file extensions [str]
    dpi: Parameter of fig.savefig()

    Accepts other arguments of fig.savefig as kwargs
    (example: bbox_inches='tight').
    """
    if "png" in formats:
        fig.savefig(filename + ".png", dpi=dpi, **kwargs)
        if confirm:
            print("Saved %s.png" % filename)
    if "pdf" in formats:
        fig.savefig(filename + ".pdf", dpi=dpi, **kwargs)
        if confirm:
            print("Saved %s.pdf" % filename)


def abs_time(s: int):
    """Return an absolute amount of seconds as hh:mm string."""
    hours = s // 3600
    minutes = (s - hours * 3600) // 60
    # seconds = s - hours * 3600 - minutes * 60
    return f"{hours:02d}:{minutes:02d}"
    # return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def as_text(value):
    """Return *value* as str, except if *value* is None, in which case an empty
    string is returned.
    """
    if value is None:
        return ""
    return str(value)


def adjust_column_width(ws):
    """Set column widths in worksheet *ws* of class
    openpyxl.worksheet.worksheet.Worksheet to max length of cell contents.
    Is exact when using a monospace font in excel (not by default), but also
    provides a decent estimation without.
    From: https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    """
    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column].width = length


def to_prev_values(vector, to_replace=None, first=0):
    """Replace values identical with *to_replace* in a list by previous value
    that is not *to_replace*. If the first value is *to_replace*, *first* is
    set as first value.
    """
    if vector:
        if vector[0] is to_replace:
            vector[0] = first
        for i, val in enumerate(vector):
            if val is to_replace:
                vector[i] = vector[i - 1]
    return vector


def discrete2continuous_logs(data_dict, first=0, subkeys=None):
    """Take a dict of discrete timestep logs and return a list with values of
    every timestep.
    *first* is the parameter for function to_prev_values.
    *subkeys* is a list of strings that may be supplied if values in
    *data_dict* are subdicts, optionally further nested. Values for the return
    list are then extracted from the subdict with all keys.
    """
    if subkeys is None:
        subkeys = []
    vector = [None] * globalConstants["general"]["SIMULATION_TIME"]
    for t, subdict in data_dict.items():
        vector[t] = reduce(operator.getitem, subkeys, subdict)

    vector = to_prev_values(vector, first=first)

    return vector


def seconds2date(si, datefmt=datefmt_general, *args):
    """Convert a second to a date based on base_date being equivalent to 0.
    Return a str of *datefmt*.
    """
    return (base_date + timedelta(seconds=si)).strftime(datefmt)


def seconds2date_major(si, *args):
    """Convert a second to a date based on base_date being equivalent to 0."""
    return seconds2date(si, datefmt_major, *args)


def seconds2date_major2(si, *args):
    """Convert a second to a date based on base_date being equivalent to 0."""
    return seconds2date(si, datefmt_major2, *args)


def seconds2date_minor(si, *args):
    """Convert a second to a date based on base_date being equivalent to 0."""
    return seconds2date(si, datefmt_minor, *args)


def to_dateaxis(ax):
    """Format a date x axis. Set base_date, datefmt and byhour at the top of
    the script. Assumes that one step equals one second.

    ax: [Axes]
    """
    fmt_major = ticker.FuncFormatter(seconds2date_major2)
    ax.xaxis.set_major_formatter(fmt_major)
    ax.xaxis.set_major_locator(ticker.MultipleLocator(xdatespacing_major))

    fmt_minor = ticker.FuncFormatter(seconds2date_minor)
    ax.xaxis.set_minor_formatter(fmt_minor)
    ax.xaxis.set_minor_locator(ticker.AutoMinorLocator(minor_intervals_per_major_tick))

    # plt.gcf().autofmt_xdate(which='both')
    ax.tick_params(axis="x", which="both", labelsize=8)
    # ax.tick_params(axis='y', labelsize=12)
    # ax.tick_params(axis='x', which='major', pad=15)
    ax.tick_params(axis="x", which="minor", pad=5)
    # ax.set_xlabel('Time', fontsize=10)
    # ax.set_xlabel('Zeit', fontsize=12)

    ax.xaxis.grid(which="major", color="grey", linewidth=1.5)
    ax.tick_params(axis="x", which="major", width=1.5)
    ax.xaxis.grid(which="minor", color="lightgrey", linestyle="--", dashes=(5, 5))
    ax.tick_params(which="minor", left=False, right=False, bottom=True)


def adjust_plt():
    plt.xlim(left=0, right=globalConstants["general"]["SIMULATION_TIME"])
    plt.ylim(bottom=0)
    plt.grid(True)
    plt.show()


def adjust_plt_hist():
    plt.xlim(left=0)
    plt.ylim(bottom=0)
    plt.gca().yaxis.grid(True)
    plt.show()


def make_patch_spines_invisible(ax):
    """Helper function for a third or following y axis.
    From: https://matplotlib.org/gallery/ticks_and_spines/multiple_yaxis_with_spines.html
    """
    ax.set_frame_on(True)
    ax.patch.set_visible(False)
    for sp in ax.spines.values():
        sp.set_visible(False)


def align_yaxis(ax1, v1, ax2, v2):
    """Adjust ax2 ylimit so that v2 in ax2 is aligned to v1 in ax1.
    Modified; from: https://stackoverflow.com/questions/10481990/matplotlib-axis-with-two-scales-shared-origin
    """
    _, y1 = ax1.transData.transform((0, v1))
    _, y2 = ax2.transData.transform((0, v2))
    inv = ax2.transData.inverted()
    _, dy = inv.transform((0, 0)) - inv.transform((0, y1 - y2))
    miny, maxy = ax2.get_ylim()
    ax2.set_ylim(miny + dy, maxy)


def calc_descr_stats(d, key):
    """Update dict *d* by some descriptive statistics based on values in list
    *d*[*key*].
    """
    intervals_array = np.asarray(d[key])
    d["min"] = intervals_array.min()
    d["max"] = intervals_array.max()
    d["mean"] = np.mean(intervals_array)
    d["median"] = np.median(intervals_array)


class Report:
    def __init__(self, depotsim):
        self.depotsim = depotsim
        self.configurator = depotsim.configurator

    def get_depot_logs(self):
        depot_logs = [None] * globalConstants["general"]["SIMULATION_TIME"]

        # Area count
        for area in itertools.chain(
            self.depotsim.depot.list_areas, [self.depotsim.depot.init_store]
        ):
            data_count = area.logger.loggedData["count"]
            key = "count " + area.ID
            for t, c in data_count.items():
                if depot_logs[t] is None:
                    depot_logs[t] = {}
                depot_logs[t][key] = c

        # Total power in depot
        key = "total power [kW]"
        for t, p in self.depotsim.depot.evaluation.power_logs.items():
            if depot_logs[t] is None:
                depot_logs[t] = {}
            depot_logs[t][key] = p

        # Fill gaps (assuming that at step 0 every info was logged once)
        all_keys = depot_logs[0].keys()
        for t, subd in enumerate(depot_logs):
            if t == 0 or depot_logs[t] is None:
                continue

            for k in all_keys:
                if k not in subd:
                    found = False
                    back = 1
                    while not found:
                        if (
                            depot_logs[t - back] is not None
                            and k in depot_logs[t - back]
                        ):
                            depot_logs[t][k] = depot_logs[t - back][k]
                            found = True
                        back += 1

        # Append time and total depot count
        keys_count = ["count " + ID for ID in self.depotsim.depot.areas.keys()] + [
            "count init"
        ]
        for t, subd in enumerate(depot_logs):
            if subd is not None:
                subd["time [s]"] = t
                subd["count depot total"] = sum([subd[a] for a in keys_count])

        # Remove entries where nothing happened
        depot_logs = [c for c in depot_logs if c is not None]

        # Convert to list of ordered lists, ready for export
        headers = ["time [s]"] + keys_count + ["total power [kW]"]
        depot_logs = [[d[header] for header in headers] for d in depot_logs]

        return depot_logs, headers

    def append_tripdata(self, wb):
        ws = wb.create_sheet("Trip Data")

        headers = [
            "ID",
            "vehicle_types",
            "distance [km]",
            "std [s]",
            "std [time]",
            "sta [s]",
            "sta [time]",
            "atd [s]",
            "atd [time]",
            "ata [s]",
            "ata [time]",
            "actual_duration [s]",
            "actual_duration [hh:mm]",
            "vehicle.ID",
            "vehicle.vehicle_type",
            "reserved_for_init",
            "vehicle_from",
            "is overdue",
            "departure delay [s]",
            "departure delay [hh:mm]",
        ]

        ws.append(headers)

        # data
        for trip in self.depotsim.depot.timetable.trips_issued:
            row = [
                trip.ID,
                trip.vehicle_types_joinedstr,
                trip.distance,
                trip.std,
                seconds2date(trip.std),
                trip.sta,
                seconds2date(trip.sta),
                trip.atd,
                seconds2date(trip.atd) if trip.atd is not None else None,
                trip.ata,
                seconds2date(trip.ata) if trip.ata is not None else None,
                trip.actual_duration,
                abs_time(trip.actual_duration)
                if trip.actual_duration is not None
                else None,
                trip.vehicle.ID if trip.vehicle is not None else None,
                trip.vehicle.vehicle_type.ID if trip.vehicle is not None else None,
                trip.reserved_for_init,
                trip.vehicle_from,
                trip.delayed_departure,
                trip.delay_departure,
                abs_time(trip.delay_departure)
                if trip.delay_departure is not None
                else None,
            ]

            ws.append(row)

        adjust_column_width(ws)

    def export_to_excel(self):
        wb = Workbook()

        ws = wb.active
        ws.title = "Vehicle Events"
        depot_logs, headers_depot = self.get_depot_logs()
        ws.append(headers_depot)
        for l in depot_logs:
            ws.append(l)
        adjust_column_width(ws)

        self.append_tripdata(wb)

        filename = self.configurator.templatename + "__results" + ".xlsx"
        wb.save(globalConstants["depot"]["path_results"] + filename)

        print(
            "Export to %s successful."
            % (globalConstants["depot"]["path_results"] + filename)
        )


class DepotAnalysis:
    """
    Class for vehicle movement analysis inside the depot. Provides excel exports.
    """

    ACTION_POP_PARK = "ACTION_POP_PARK"
    ACTION_POP_UNPARK = "ACTION_POP_UNPARK"
    ACTION_POP_DEPART = "ACTION_POP_DEPART"
    ACTION_CHARGE_START = "ACTION_CHARGE_START"
    ACTION_CHARGE_FULL = "ACTION_CHARGE_FULL"
    ACTION_PROCESS_FINISHED = "ACTION_PROCESS_FINISHED"
    ACTION_PROCESS_CALLED = "ACTION_PROCESS_CALLED"

    def __init__(self, depotsim):
        self.depotsim = depotsim
        self.configurator = depotsim.configurator
        self.depot = depotsim.depot
        self.env = depotsim.env
        self.logs = []

    @property
    def defaultname(self):
        """Return a filename for export excluding file extension."""
        return (
            self.configurator.templatename
            + "_export_"
            + self.depotsim.evaluation.now_repr
        )

    def log(self, event=None):
        if isinstance(event, StorePut):
            eventname = self.ACTION_POP_PARK
        elif isinstance(event, StoreGet):
            eventname = self.ACTION_POP_UNPARK
        elif isinstance(event, Departure):
            eventname = self.ACTION_POP_DEPART
        elif isinstance(event, FullyCharged):
            eventname = self.ACTION_CHARGE_FULL
        elif isinstance(event, ChargeStart):
            eventname = self.ACTION_CHARGE_START
        elif isinstance(event, ProcessFinished):
            eventname = self.ACTION_PROCESS_FINISHED
        elif isinstance(event, ProcessCalled):
            eventname = self.ACTION_PROCESS_CALLED
        else:
            raise ValueError("Unknown event %s" % event)

        log = DepotLog(self.env.now, eventname, event)
        self.logs.append(log)

        log.area_counts = {}
        log.depot_count_total = 0

        for area in self.depot.list_areas:
            area_population = area.count
            log.area_counts[area.ID] = area_population
            log.depot_count_total += area_population

        log.init_store_count = self.depot.init_store.count
        log.overdue_trips = len(self.depot.overdue_trips)

    def export_logs(self, filename):
        """
        Exports the list of logs into an Excel file located at [filename].

        filename: [str] excluding path and file extension
        """
        if not self.logs:
            raise RuntimeError(
                "Cannot export because logging of required data is switched "
                "off in globalConstants['depot']['log_cm_data']. Rerun the "
                "simulation with the option switched on."
            )

        if filename:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "VehicleEvents"

                # meta
                meta_dateTime = ["time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
                ws.append(meta_dateTime)

                ws.append(["templatename", self.configurator.templatename])
                ws.append(["depot ID", self.depot.ID])

                ws.append([])  # empty row
                ws.append([])  # empty row

                # headers
                headers = [
                    "event ID",
                    "sim time",
                    "vehicle ID",
                    "vehicle type",
                    "battery level",
                    "battery capacity",
                    "area ID",
                    "power",
                    "event name",
                    "event relevance CM",
                    "overdue trips",
                    "depot total count",
                ]

                for area in self.depot.areas.values():
                    headers.append(area.ID + " count")
                headers.append(self.depot.init_store.ID + " count")

                ws.append(headers)

                # data
                for event_ID, log in enumerate(self.logs):
                    current_area_ID = log.area.ID if log.area is not None else ""
                    power = None
                    if log.action == "ACTION_CHARGE_FULL":
                        power = log.area.charging_interfaces[log.slot].max_power

                    # action relevance check
                    relevance_cm = (
                        log.area is not None
                        and log.area.issink
                        and (
                            log.action == "ACTION_CHARGE_START"
                            or log.action == "ACTION_CHARGE_FULL"
                            or log.action == "ACTION_POP_DEPART"
                        )
                    )

                    row = [
                        event_ID,
                        log.simTime,
                        log.vehicle.ID,
                        log.vehicle.vehicle_type.ID,
                        log.battery_level,
                        log.vehicle.battery.energy_real,
                        current_area_ID,
                        power,
                        log.action,
                        relevance_cm,
                        log.overdue_trips,
                        log.depot_count_total,
                    ]

                    for area_ID in log.area_counts:
                        row.append(log.area_counts[area_ID])

                    row.append(log.init_store_count)

                    if len(row) > 0:
                        ws.append(row)

                # Add another sheet with trip data
                self.append_tripdata(wb)

                # save
                if hasattr(self.depot, "view"):
                    wb.save(self.depot.view.path_results + filename + ".xlsx")
                    return True
                else:
                    filename_full = (
                        globalConstants["depot"]["path_results"] + filename + ".xlsx"
                    )
                    wb.save(filename_full)
                    print(
                        "Vehicle event data and trip data exported to %s"
                        % filename_full
                    )

            except:
                traceback.print_exc()
                if hasattr(self.depot, "view"):
                    return False
                else:
                    raise RuntimeError("Error while exporting depot analysis!")

    def export_logs_smart_charging(self):
        """
        Exports the list of logs into pd_frame.

        :return: pd_frame
        """
        if not self.logs:
            raise RuntimeError(
                "Cannot export because logging of required data is switched "
                "off in globalConstants['depot']['log_cm_data']. Rerun the "
                "simulation with the option switched on."
            )
        # headers
        headers = [
            "sim_time",
            "vehicle_ID",
            "vehicle_type",
            "battery_level",
            "battery_capacity",
            "power",
            "event_name",
        ]

        return_frame = pd.DataFrame(columns=headers)

        # data
        for event_ID, log in enumerate(self.logs):
            power = None
            if log.action == "ACTION_CHARGE_FULL":
                power = log.area.charging_interfaces[log.slot].max_power

            # action relevance check
            relevance_cm = log.area is not None and (
                log.action == "ACTION_CHARGE_START"
                or log.action == "ACTION_CHARGE_FULL"
                or log.action == "ACTION_POP_DEPART"
            )
            if relevance_cm:
                row = [
                    log.simTime,
                    log.vehicle.ID,
                    log.vehicle.vehicle_type.ID,
                    log.battery_level,
                    log.vehicle.battery.energy_real,
                    power,
                    log.action,
                ]
                return_frame.loc[event_ID] = row

        return return_frame

    def append_tripdata(self, wb):
        """
        Extend wb by trip data.
        """
        # wb = Workbook()
        ws = wb.create_sheet("TripData")
        # ws = wb.active

        # headers
        headers = [
            "ID",
            "origin",
            "destination",
            "vehicle_types",
            "std",
            "sta",
            "distance",
            "atd",
            "ata",
            "eta",
            "vehicle.ID",
            "vehicle.vehicleType",
            "reserved_for_init",
            "vehicle_from",
            "overdue",
        ]

        ws.append(headers)

        # Add sheet with summarizing infos
        ws2 = wb.create_sheet("TripChecks")
        vehicle_types = list(globalConstants["depot"]["vehicle_types"].keys())
        overdueChecks = []
        # overdueByType = {vehicleType: 0 for vehicleType in
        #                 ['EN', 'GN', 'DL']}
        overdueByType = {vehicleType: 0 for vehicleType in vehicle_types}
        uniqueVehicles = []
        # vehicleCount = {vehicleType: 0 for vehicleType in
        #                 ['EN', 'GN', 'DL']}
        vehicleCount = {vehicleType: 0 for vehicleType in vehicle_types}

        # data
        for trip in self.depot.timetable.trips_issued:
            pos1 = trip.vehicle.ID if trip.vehicle is not None else None
            pos2 = trip.vehicle.vehicle_type.ID if trip.vehicle is not None else None
            pos3 = trip.vehicle_from if hasattr(trip, "vehicle_from") else None
            pos4 = trip.delayed_departure

            row = [
                trip.ID,
                trip.origin.ID,
                trip.destination.ID,
                trip.vehicle_types_joinedstr,
                trip.std,
                trip.sta,
                trip.distance,
                trip.atd,
                trip.ata,
                trip.eta,
                pos1,
                pos2,
                trip.reserved_for_init,
                pos3,
                pos4,
            ]

            if len(row) > 0:
                ws.append(row)

            # For summarizing info
            if pos4 is not None:
                overdueChecks.append(pos4)
                if pos4:
                    overdueByType[trip.vehicle_types_joinedstr] += 1
            if trip.vehicle is not None:
                if trip.vehicle.ID not in uniqueVehicles:
                    uniqueVehicles.append(trip.vehicle.ID)
                    vehicleCount[trip.vehicle.vehicle_type.ID] += 1

        overdueCheck = any(overdueChecks)

        # Get additional data from init_store
        occurrences = [
            vehicle.vehicle_type.ID for vehicle in self.depot.init_store.items
        ]
        typeCounter = Counter(occurrences)
        tck = list(typeCounter.keys())
        tck = ["VehicleStoreInit " + k + " at end" for k in tck]

        # Summarize
        headers2 = [
            "any overdue",
            "EN overdues",
            "GN overdues",
            "DL overdues",
            "EN unique",
            "GN unique",
            "DL unique",
        ] + tck

        addData = (
            [overdueCheck]
            + list(overdueByType.values())
            + list(vehicleCount.values())
            + list(typeCounter.values())
        )

        # Fill sheet (transposed)
        for ti, va in zip(headers2, addData):
            ws2.append([ti, va])
        # ws2.append(headers2)
        # ws2.append(addData)


class DepotLog:
    def __init__(self, simTime, action, event):
        """
        simTime: simpy environment simTime
        vehicle: eflips.SimpleVehicle instance
        action: string, available options in [DepotAnalysis]
        event: simpy put or get event or FullyCharged or Departure.
            Must be provided if *action* is ACTION_POP_PARK or
            ACTION_POP_UNPARK.
        """
        self.simTime = simTime
        self.action = action
        self.slot = None

        if action == "ACTION_POP_PARK":
            self.vehicle = event.item
            self.area = event.resource
        elif action == "ACTION_POP_UNPARK":
            self.vehicle = event.value
            self.area = event.resource
        else:
            self.vehicle = event.item
            self.area = self.vehicle.dwd.current_area

        if action == "ACTION_CHARGE_FULL" or action == "ACTION_CHARGE_START":
            self.slot = self.vehicle.dwd.current_area.items.index(self.vehicle)

        self.battery_level = self.vehicle.battery.energy


class ChargeStart:
    eventname = "ACTION_CHARGE_START"

    def __init__(self, env, item):
        self.env = env
        self.item = item


class FullyCharged:
    eventname = "ACTION_CHARGE_FULL"

    def __init__(self, env, item):
        self.env = env
        self.item = item


class ProcessFinished:
    eventname = "ACTION_PROCESS_FINISHED"

    def __init__(self, env, item):
        self.env = env
        self.item = item


class ProcessCalled:
    """Lo for time, when vehicle would like to proceed to next area. Needed for congestion calculation."""

    eventname = "ACTION_PROCESS_CALLED"

    def __init__(self, env, item):
        self.env = env
        self.item = item


class Departure:
    eventname = "ACTION_POP_DEPART"

    def __init__(self, env, item):
        self.env = env
        self.item = item
