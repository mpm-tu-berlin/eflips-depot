# -*- coding: utf-8 -*-
"""
Created on Thu Sep 28 16:31:00 2017

@author: T.Altmann; P. Boev

Version 1.1 (PM, 29.09.2017):
    - log() overhaul, including bugfix for event logging.
    - Added get_valList() for getting a plottable list from loggedData-dict.
    - Enabled passing string to log() that represents multilevel access to 
        attributes of sub-objects of logObj. E.g. attr = 'vehicle.ID'
Version 1.2 (PB, 12.12.17)
    - Added data gatherer class for collection of the logged data
"""
from functools import reduce
from math import ceil
import os
import numpy as np
import numbers
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib
import xlsxwriter
import copy
import pandas as pd
import weakref
from eflips.settings import globalConstants, rcParams
from eflips.helperFunctions import createEvalScheme, add_class_attribute, \
    complex_setter, change_getter
from inspect import getsourcelines
from openpyxl import load_workbook
import matplotlib.patches as patches
import math
from itertools import cycle

# Caution: rcParams is not actively called, but required for matplotlib

class DataLogger:
    """Log attributes and properties of one object continuously or only
    at event occurrence.
    A DataLogger object is bound to one logObj. Therefore initialize one
    DataLogger for each logObj.

    attsToLog: [list] of strings specifying what attributes and properties to
        log continuously (at every (!) timestep). If logging is supposed to
        happen only at event occurrence, then exclude it from attsToLog and
        manually call log() where the event occurs.
    loggedData: [dict] container for all logged data, to be accessed after the
        simulation. Object-specific dict with attribute/property names as keys
        for sub-dicts that contain timestep-value pairs; Example:
            loggedData =    {'nVechicles': {
                                0: 0,
                                1: 0,
                                2: 1,
                                ...},
                             'powerOutput': {
                                0: 0,
                                1: 0,
                                2: 50,
                                ...}}
        Reasons for using sub-dicts over simple lists: 
            1. enable event logging where not every timestep is documented, and 
                therefore the timestep needs to be logged as well
            2. enable safely getting timestep-specific values

    classToLog has to be existing key in the evaluation scheme in settings
    """

    def __init__(self, env, logObj, classToLog):
        self.env = env
        self.logObj = logObj
        self.attsToLog = createEvalScheme(classToLog)
        self.check_attsToLog()
        self.evaluationSets = globalConstants['evaluationScheme'][classToLog]
        self.loggedData = {}
        self.action = None

        self.steplogSwitch = globalConstants['general']['LOG_ATTRIBUTES'] and any(
            self.attsToLog['attsToLog_time']) and globalConstants['general'][
            'LOG_SPECIFIC_STEPS']   # evaluate here because it's constant

        # Event logging for attributes that
        # need to be logged during object initialization
        if globalConstants['general']['LOG_ATTRIBUTES'] and \
                any(self.attsToLog['attsToLog_time']):
            for attr in self.attsToLog['attsToLog_const']:
                self.log(attr)

        # log at specific events (first time point)
        if globalConstants['general']['LOG_ATTRIBUTES'] and\
                any(self.attsToLog['attsToLog_time']) and\
                globalConstants['general']['LOG_SPECIFIC_STEPS'] and\
                globalConstants['general']['LOG_CONTINUOUSLY'] is False:
            for attr in self.attsToLog['attsToLog_time']:
                self.log(attr)

        # log at every timestep
        if globalConstants['general']['LOG_ATTRIBUTES'] and \
                any(self.attsToLog['attsToLog_time']) and \
                globalConstants['general']['LOG_CONTINUOUSLY']:
            self.action = env.process(self.run())

        # log with modified properties (getter and setter)
        if globalConstants['general']['LOG_ATTRIBUTES'] and \
                any(self.attsToLog['attsToLog_time']) and \
                globalConstants['general']['LOG_SPECIFIC_STEPS'] is False and \
                globalConstants['general']['LOG_CONTINUOUSLY'] is False:
            # Event logging at object initialization for attributes that
            # need to be logged over the simulation time
            for attr in self.attsToLog['attsToLog_time']:
                self.log(attr)
            if hasattr(self.logObj, 'modified') is False:
                # add instances attribute
                self.logObj.__class__.instances = list()
                self.logObj.instances.append(weakref.ref(self.logObj))
                for attr in self.attsToLog['attsToLog_time']:
                    self.patch_attr(self.logObj, attr)
                setattr(self.logObj.__class__, 'modified', True)
            else:
                for attr in self.attsToLog['attsToLog_time']:
                    self.extendObj(self.logObj, attr)
            # self.action = simpy.Event(env)
            # self.action._ok = True
            # self.action._value = self
            # self.action.callbacks.append(self.transfer_logged_atts)
            # self.env.schedule(self.action, URGENT,
            #                   globalConstants['general']['SIMULATION_TIME']
            #                   - self.env.now)
            self.env.process(self.transfer_logged_atts())

    def run(self):
        """Call log() for each attr in attsToLog at every timestep"""
        while True:
            for attr in self.attsToLog['attsToLog_time']:
                self.log(attr)
            yield self.env.timeout(1)

    def steplog(self, *args, **kwargs):
        """Call log() for each attr in attsToLog at specific
        timestep in code.
        Accepts *args/**kwargs in order to allow calling this method as a simpy
        event callback.
        """
        if self.steplogSwitch:
            for attr in self.attsToLog['attsToLog_time']:
                self.log(attr)

    def patch_attr(self, mainObject, attr):
        """modifies the class which has an attribute to log

        Pavel Boev
        """
        sub_attr_list = attr.split('.')
        # attribute in main object (vehicle.attr)
        if len(sub_attr_list) == 1:
            # check if attribute is a property; if not:
            if eval('isinstance(getattr(type(self.logObj)'
                    + ', \'' + attr + '\', None)'
                    + ', property)') is False:
                # if not a property add attribute with '_'
                new_attr_name = '_' + attr
                add_class_attribute(mainObject, attr,
                                    new_attr_name)
                # set the property with getter and setter;
                # works only on classes and not instances
                setattr(mainObject.__class__, attr,
                        property(lambda self:
                                 getattr(self, new_attr_name),
                                 lambda self, value:
                                 complex_setter(self, mainObject.__class__,
                                                attr,
                                                new_attr_name,
                                                value)))
            else:  # attribute is a property
                class_attribute = eval('getattr(type(self.logObj)'
                                       + ', \''
                                       + attr
                                       + '\', None)')
                # check if property has only getter
                if class_attribute.fset is None:
                    # get what the getter returns and
                    # expand it with a logger
                    funtext = getsourcelines(class_attribute.fget)
                    last_line = funtext[0][-1]
                    last_line.strip()
                    fun_str = \
                        last_line[last_line.find('return') + 7:]
                    setattr(mainObject.__class__, attr,
                            property(lambda self:
                                     change_getter(self, mainObject.__class__,
                                                   attr,
                                                   fun_str)))
                else:  # property has also a setter
                    # change getter and setter
                    new_attr_name = '_' + attr
                    setattr(mainObject.__class__, attr,
                            property(lambda self:
                                     getattr(self, attr),
                                     lambda self, value:
                                     complex_setter(self, mainObject.__class__,
                                                    attr,
                                                    new_attr_name,
                                                    value)))
        # attribute in nested object (e.g. vehicle.object.attr)
        else:
            sub_attr_name = sub_attr_list[-1]
            logObj_name = 'self.logObj'
            for i in range(len(sub_attr_list) - 1):
                logObj_name += '.' + sub_attr_list[i]
            logObj_name_class = logObj_name + '.__class__'
            if hasattr(eval(logObj_name), 'logger') is False:
                setattr(eval(logObj_name), 'logger', {})
            # check if attribute is a property; if not:
            if eval('isinstance(getattr(type(' + logObj_name + ')'
                    + ', \'' + sub_attr_name + '\', None)'
                    + ', property)') is False:
                # if not a property add attribute with '_'
                new_attr_name = '_' + sub_attr_name
                add_class_attribute(eval(logObj_name),
                                    sub_attr_name,
                                    new_attr_name)

                # set the property with getter and setter;
                # works only on classes and not instances
                setattr(eval(logObj_name_class), sub_attr_name,
                        property(lambda self:
                                 getattr(self, new_attr_name),
                                 lambda self, value:
                                 complex_setter(self, mainObject.__class__,
                                                sub_attr_name,
                                                new_attr_name,
                                                value)))
            else:  # attribute is a property
                class_attribute = eval(
                    'getattr(type(' + logObj_name + ')'
                    + ', \''
                    + sub_attr_name
                    + '\', None)')
                # check if property has only getter
                if class_attribute.fset is None:
                    # get what the getter returns and
                    # expand it with a logger
                    funtext = getsourcelines(class_attribute.fget)
                    last_line = funtext[0][-1]
                    last_line.strip()
                    fun_str = \
                        last_line[last_line.find('return') + 7:]
                    setattr(eval(logObj_name_class), sub_attr_name,
                            property(lambda self:
                                     change_getter(self, mainObject.__class__,
                                                   sub_attr_name,
                                                   fun_str)))
                else:  # property has also a setter
                    # change getter and setter
                    new_attr_name = '_' + sub_attr_name
                    setattr(eval(logObj_name_class), sub_attr_name,
                            property(lambda self:
                                     getattr(self, new_attr_name),
                                     lambda self, value:
                                     complex_setter(self, mainObject.__class__,
                                                    sub_attr_name,
                                                    new_attr_name,
                                                    value)))

    def extendObj(self, mainObject, attr):
        """extends logged object attributes with _attribute and adds
        internal object logger, only if class already modified

        Pavel Boev
        """
        sub_attr_list = attr.split('.')
        # attribute in main object (vehicle.attr)
        if len(sub_attr_list) == 1:
            # check if attribute is a property; if not:
            if eval('isinstance(getattr(type(self.logObj)'
                    + ', \'' + attr + '\', None)'
                    + ', property)') is False:
                # if not a property add attribute with '_'
                new_attr_name = '_' + attr
                add_class_attribute(mainObject, attr,
                                    new_attr_name)
        # attribute in nested object (e.g. vehicle.object.attr)
        else:
            sub_attr_name = sub_attr_list[-1]
            logObj_name = 'self.logObj'
            for i in range(len(sub_attr_list) - 1):
                logObj_name += '.' + sub_attr_list[i]
            if hasattr(eval(logObj_name),
                       'logger') is False:
                setattr(eval(logObj_name), 'logger', {})
            # check if attribute is a property; if not:
            if eval('isinstance(getattr(type(' + logObj_name + ')'
                    + ', \'' + sub_attr_name + '\', None)'
                    + ', property)') is False:
                # if not a property add attribute with '_'
                new_attr_name = '_' + sub_attr_name
                add_class_attribute(eval(logObj_name),
                                    sub_attr_name,
                                    new_attr_name)

    def transfer_logged_atts(self):
        """this function transfers the logged data from a nested logged
        to the main object logger

        Pavel Boev
        """
        yield self.env.timeout(globalConstants['general']['SIMULATION_TIME'] -
                               self.env.now - 1)
        for attr in self.attsToLog['attsToLog_time']:
            sub_attr_list = attr.split('.')
            if len(sub_attr_list) > 1:
                loggedDataPath = 'self.logObj'
                for part in sub_attr_list[:-1]:
                    loggedDataPath += '.' + part
                if hasattr(eval(loggedDataPath), 'logger'):
                    if sub_attr_list[-1] in eval(loggedDataPath).logger:
                        d = eval(loggedDataPath).logger[sub_attr_list[-1]]
                        self.logObj.logger.loggedData[attr].update(d)
                        del eval(loggedDataPath).logger[sub_attr_list[-1]]

    def log(self, attr):
        """Get *attr* and save it to loggedData.
        attr: [str] name of the attribute/property to log.
            Can handle direct attribute arguments (e.g. attr = 'name') as well
            as multilevel argument as a single string (e.g. attr =
            'battery.soc').
            attr must be accessible from self.logObj. Otherwise, None will be
            logged.
        """
        if globalConstants['general']['LOG_ATTRIBUTES']:
            subDict = self.loggedData.get(attr, False)
            # create an empty dict for attr if it's its first log
            if not subDict:
                self.loggedData[attr] = {}

            try:
                value = reduce(getattr, [self.logObj] + attr.split('.'))
            except AttributeError:
                # Log None in case the attribute doesn't exist.
                # Enables continuous logging of attributes and objects
                # that are nonexistent during parts of the simulation.
                self.loggedData[attr][int(self.env.now)] = None
            else:
                # no AttributeError, go on
                if globalConstants['general']['LOG_COPIES']:
                    self.loggedData[attr][self.env.now] = copy.deepcopy(value)
                else:
                    self.loggedData[attr][int(self.env.now)] = value

    def get_valList(self, attr, SIM_TIME=False):
        """Return a list *y* with all values from the dict loggedData[attr].
        Purpose: getting a plottable vector.
        Works for data that was logged at every timestep as well as data logged
        only at events. Returns None values for timesteps where no data was
        logged.

        attr: [str] attribute/property ID. If an attribute from a sub-object
        SIM_TIME: [int] total simulation time. Determines the length of y.
        If no SIM_TIME is passed, it is assumed to be the maximum
        logged timestep. Therefore y might be shorter than
        the actual SIM_TIME if the maximum logged timestep
        is not equal to SIM_TIME.
        """
        if not SIM_TIME:
            SIM_TIME = ceil(max(self.loggedData[attr].keys()) + 1)
        y = [None] * SIM_TIME
        for key in self.loggedData[attr]:
            y[key] = self.loggedData[attr][key]
        #        x = list(range(y[0], y[-1] + 1))
        #        return x, y
        return y

    def check_attsToLog(self):
        for attr in self.attsToLog['attsToLog_time']:
            if not isinstance(attr, str):
                raise ValueError(
                    "list attsToLog for class '%s' " % (type(self.logObj))
                    + "must only contain strings")
        return True


class DataGatherer:
    """
    Gathers the logged data from the DataLogger
    e.g. bus_data = DataGatherer(vehicles)
    can be used for all objects or for one only (e.g. vehicles[1])

    The logged data is saved in the dict .data.
    Unzip(), unzip_all() and deep_unzip()
    structure the logged data from one attribute as plottable data in dict
    """

    def __init__(self, loggedObj):
        self.loggedObj = loggedObj
        self.data = {}
        self.attsToLog_initial = []
        self.attsToLog = []
        self.evaluationSets = []
        self.globalConstants = globalConstants

        if globalConstants['general']['LOG_ATTRIBUTES'] is False:
            print('No logged data available!')
        else:
            if isinstance(self.loggedObj, list):
                if hasattr(self.loggedObj[0], 'logger') is False:
                    print('No logged data available!')
                    return
            else:
                if hasattr(self.loggedObj, 'logger') is False:
                    print('No logged data available!')
                    return
            self.run()
            self.unzip_all()

    def run(self):
        """
        collects the logged data from the logged object (one or many) and
        saves it.
        e.g vehicles or vehicles[1]
        """
        # for a list of objects an index is added
        if isinstance(self.loggedObj, list):
            self.evaluationSets = self.loggedObj[0].logger.evaluationSets
            for idx, ID in enumerate(self.loggedObj):
                self.data[idx] = {}
                attributes = \
                    self.loggedObj[idx].logger.attsToLog['attsToLog_time'] \
                    + self.loggedObj[idx].logger.attsToLog['attsToLog_const']
                self.attsToLog_initial = attributes
                for attribute in attributes:
                    self.data[idx][attribute] = \
                        self.loggedObj[idx].logger.loggedData[attribute]
        else:  # if isinstance(self.loggedObj, list); if single object
            self.evaluationSets = self.loggedObj.logger.evaluationSets
            self.data[0] = {}
            attributes = self.loggedObj.logger.attsToLog['attsToLog_time'] \
                         + self.loggedObj.logger.attsToLog['attsToLog_const']
            self.attsToLog_initial = attributes
            for attribute in attributes:
                self.data[0][attribute] = \
                    self.loggedObj.logger.loggedData[attribute]

    # ----------------------------------------------------------------------
    # UNPACK DATA
    # ----------------------------------------------------------------------

    def unzip(self, attr):
        """
        function to transform dict of objects with all attributes to dict of
        plottable values:
        e.g. data[n]['state'] is object of type VehicleState -->
        data[n]['state.ignitionOn'];
        data[n][state.velocity]; ....
        """
        dict_attr_obj = []
        for idx, ID in enumerate(list(self.data.keys())):
            first_not_None = next((el for el in (self.data[idx][attr]).values()
                                   if el is not None), None)
            # if all are None, transform to numpy.nan
            if first_not_None is None:
                for key in self.data[idx][attr]:
                    self.data[idx][attr][key] = np.nan
            # if first_not_None is None; if there are not None objects
            else:
                # unzipping only objects and not
                # numbers or strings or dicts or lists
                if isinstance(first_not_None, numbers.Number) or \
                        isinstance(first_not_None, str) or \
                        isinstance(first_not_None, dict) or \
                        isinstance(first_not_None, list):
                    # helper variable for the other unzipping functions
                    dict_attr_obj.append(True)
                    # print(attr, ': Nothing to unzip.
                    # Attribute is good to go.')
                    if attr not in self.attsToLog:
                        self.attsToLog.append(attr)
                else:  # run only if dict of Objects
                    dict_attr_obj.append(False)
                    # replace first not none with list of all attribute
                    # of all classes (attrs of gridpoint and gridsegment )
                    attributes = []
                    for key in self.data[idx][attr]:
                        if self.data[idx][attr][key] is not None:
                            attributes = attributes \
                                         + list(
                                                set(vars(self.data[idx][attr]
                                                         [key]).keys())
                                                - set(attributes))
                    for sub_attr in attributes:
                        if sub_attr != 'env':  # no need for env
                            new_attr = attr + '.' + sub_attr
                            # create new dict e.g. data[1]['state.velocity']
                            self.data[idx][new_attr] = {}
                            for idx_time in self.data[idx][attr].keys():
                                if self.data[idx][attr][idx_time] \
                                        is not None:
                                    if hasattr(self.data[idx][attr][idx_time],
                                               sub_attr):
                                        self.data[idx][new_attr][idx_time] = \
                                            vars(self.data[idx][attr][
                                                     idx_time]).get(sub_attr)
                                    # if sub_attr doesn't exist, add None
                                    else:
                                        self.data[idx][new_attr][idx_time] = \
                                            None
                                else:  # else is None
                                    self.data[idx][new_attr][idx_time] = \
                                        None
                    # remove attr after unzipping content
                    self.data[idx].pop(attr, None)
                    # print(attr, ': Done.')
            for k in list(self.data[idx].keys()):
                if k not in self.attsToLog:
                    self.attsToLog.append(k)
        return all(dict_attr_obj)

    def unzip_single(self, attr):
        """
        unzips all layers of a single logged object:
        e.g.: if loggedvehicle.data[1]['state'] is dict of all state objects
        unzip_single('state') -->
        adds state._payload, state.slope, state.velocity, state.location.ID,
        state.location.name, state.location.type to
        loggedvehicle.data[1].keys()
        """
        # help list; stops the while loop when all elements are False
        help_attr_bool = []
        help_var = self.unzip(attr)
        help_attr_bool.append(help_var)
        # if there is a false statement in the list do:
        while all(help_attr_bool) is False:
            help_attr_bool = []
            attributes = list(self.data[0].keys())  # first loop of attributes
            matching_attributes = [s for s in attributes if attr in s]
            for matching_attribute in matching_attributes:
                help_var = self.unzip(matching_attribute)
                help_attr_bool.append(help_var)

    def unzip_all(self):
        """
        unzips all dict keys to the last layer
        """
        print('--------------------------------')
        print('Data Gatherer collecting data...')
        attributes = list(self.data[0].keys())  # first loop of attributes
        for attr in attributes:
            self.unzip_single(attr)
        print('Done.')

    def plot(self, attr, ID=None, **kwargs):
        """
        plot('state.velocity', 1)
        if ID left empty, function plots all available IDs
        """
        if attr in self.attsToLog:
            if ID is None:
                for idx, ID in enumerate(list(self.data.keys())):
                    plt.plot(*zip(*sorted(self.data[ID][attr].items())),
                             label='ID: %s' % str(ID), **kwargs)
                plt.xlabel('time (s)')
                # if '.' in attr remove content prior to the dot
                if '.' in attr:
                    attr = attr.split('.')[-1]
                plt.ylabel(attr)
                plt.title('plot of ' + attr)
                plt.grid(True)
                plt.legend()
                plt.show()
            else:
                plt.plot(*zip(*sorted(self.data[ID][attr].items())),
                         label='ID: %s' % str(ID), **kwargs)
                plt.xlabel('time (s)')
                # if '.' in attr remove content prior to the dot
                if '.' in attr:
                    attr = attr.split('.')[-1]
                plt.ylabel(attr)
                plt.title('plot of ' + attr)
                plt.grid(True)
                plt.legend()
                plt.show()
        else:
            print('Cannot find attribute "%s"' % attr)

    def save(self, path, basename, append_date=False):
        """convert gathered data to pandas Dataframe and
        export to hdf5
        """
        if bool(self.data) is False:
            print('DataGatherer is empty. No file saved.')
        else:
            self.save_data = {}
            for i in self.data:
                loggedAtts_const = {}
                loggedAtts_time = {}
                for k, v in self.data[i].items():
                    if len(self.data[i][k]) == 1:
                        firstKey = list(self.data[i][k].keys())[0]
                        loggedAtts_const.update({k: self.data[i][k][firstKey]})
                    else:  # save dict as pandas TimeSeries
                        loggedAtts_time.update(
                            {k: pd.Series(list(self.data[i][k].values()),
                                          pd.to_datetime(
                                              list(self.data[i][k].keys()),
                                              unit='s',
                                              origin=
                                              pd.Timestamp('2018-01-01')))})
                self.save_data[i] = {}
                self.save_data[i]['loggedAtts_const'] = loggedAtts_const
                self.save_data[i]['loggedAtts_time'] = loggedAtts_time

            d = {'data': self.save_data,
                 'globalConstants': self.globalConstants,
                 'evaluationSets': self.evaluationSets}
            if append_date:
                filename_save = path + '\\' + basename + '_' +\
                                globalConstants['EXEC_TIME']\
                                + '.h5'
            else:
                filename_save = path + '\\' + basename + '.h5'
            dd.io.save(filename_save, d)
            print('Data saved to file "' + filename_save + "\"")


###############################################################################
# EVALUATION CLASS
###############################################################################


class Evaluation:
    """
    evaluate and plot data
    input is from DataGatherer exported hdf5 file
    if filename is left empty, a file explorer pops up for selecting a data
     file
    """

    def __init__(self, path=None, filename=None):
        if filename is None:
            # we don't want a full GUI, so keep the root window from appearing
            Tk().withdraw()
            # show an "Open" dialog box and
            # return the path to the selected file
            filename = filedialog.askopenfilename(
                title="Select HDF5 file",
                filetypes=(('HDF5 files', '*.hdf5;*.h5'), ('all files', '.*')))
            self.filename = filename
        else:
            self.filename = path + '\\' + filename + '.h5' if path is not None\
                else filename + '.h5'
        self.data = dd.io.load(self.filename, '/data')
        self.globalConstants = dd.io.load(self.filename, '/globalConstants')
        self.evaluationSets = dd.io.load(self.filename, '/evaluationSets')
        self.loggedAtts = list(self.data[0]['loggedAtts_const'].keys()) \
                          + list(self.data[0]['loggedAtts_time'])
        self.filename = os.path.basename(self.filename)
        self.calculatedAtts = []
        for ID in self.data.keys():
            self.data[ID]['calculatedAtts_const'] = {}
            self.data[ID]['calculatedAtts_time'] = {}
        print('--------------------------------')
        print('Data from file "' + self.filename + '" loaded.')

    def plot_numVehiclesCharging(self, IDs=None, show=False, save=True,
                                 path='.', baseFilename='numVehiclesCharging_',
                                 figureSize=None, formats=None):
        """Plot the number of vehicles present in each charging facility.
        Will be moved to EvaluationInfra or so later.

        Dominic Jefferies"""
        if not IDs:
            # plot for all facilities
            IDs = self.data.keys()

        for ID in IDs:
            # Collect data
            currentData = self.data[ID]
            plot_title = currentData['loggedAtts_const']['location.name']
            ydata = currentData['loggedAtts_time']['numVehicles']
            # We need this weird hack for pyplot.step():
            xdata = ydata.index.values

            plot = LinePlot(title=plot_title, xlabel='Time (hours)',
                            ylabel='Number of vehicles',
                            timelabel='hours', figureSize=figureSize,
                            show=show)
            plot.addSeries(xdata, ydata, step=True)
            # ymax = max(plot.axes.get_yticks())
            # yticks = list(range(0,ymax))
            # plot.axes.set(yticks=yticks)
            plot.axes.yaxis.set_major_locator(matplotlib.ticker.
                                              MultipleLocator(base=1))

            if save:
                plot_filename = baseFilename + '{:04d}'.format(ID)
                plot.save(path, plot_filename, formats=formats)

    def plot_maxPower(self, IDs=None, show=False, save=True,
                      path='.', baseFilename='maxPower_',
                      formats=None):
        """Plot the number of vehicles present in each charging facility.
        Will be moved to EvaluationInfra or so later.

        Dominic Jefferies"""
        if not IDs:
            # plot for all facilities
            IDs = self.data.keys()

        for ID in IDs:
            # Collect data
            currentData = self.data[ID]
            plot_title = currentData['loggedAtts_const']['location.name']
            ydata = currentData['loggedAtts_time']['numVehicles'] * \
                    currentData['loggedAtts_const']['interface.maxPower'] / \
                    currentData['loggedAtts_const']['interface.efficiency']
            # We need this weird hack for pyplot.step():
            xdata = ydata.index.values

            plot = LinePlot(title=plot_title, xlabel='Time (hours)',
                            ylabel='Power demand (kW)',
                            timelabel='hours', show=show)
            plot.addSeries(xdata, ydata, step=True)
            # ymax = max(plot.axes.get_yticks())
            # yticks = list(range(0,ymax))
            # plot.axes.set(yticks=yticks)
            # plot.axes.yaxis.set_major_locator(matplotlib.ticker.\
            #                                   MultipleLocator(base=1))

            if save:
                plot_filename = baseFilename + '{:04d}'.format(ID)
                plot.save(path, plot_filename, formats=formats)

    def plot_slotOccupation(self, IDs=None, show=False, save=True,
                            path='.', baseFilename='slotOccupation_',
                            formats=False):
        """"""
        if not IDs:
            # plot for all facilities
            IDs = self.data.keys()

        for ID in IDs:
            currentData = self.data[ID]
            plot_title = currentData['loggedAtts_const']['location.name']
            plot = LinePlot(title=plot_title, xlabel='Time (hours)',
                            ylabel='Slot (ID)',
                            timelabel='hours', show=show, lineWidth=2.0)
            numSlots = currentData['loggedAtts_const']['numSlots']
            occupied = currentData['loggedAtts_time']['slots.occupied']
            xdata = occupied.index.values
            colors = iter(cycle(['g', 'c', 'r', 'b', 'm', 'y']))
            for slotID in range(1,numSlots+1):
                ydata = pd.Series(data = np.array(list(map(lambda x: slotID if x and slotID in x else np.nan, occupied.values))),
                                  index = xdata)
                color = next(colors)
                plot.addRectangles(xdata, ydata, color=color)

            if save:
                plot_filename = baseFilename + '{:04d}'.format(ID)
                plot.save(path, plot_filename, formats=formats)

class EvaluationVehicle(Evaluation):
    """
    Class that evaluates all attributes from the evaluationSets defined in
    "settings.py" for all vehicle-class-objects.
    Pascal Weigmann
    """

    def __init__(self, path=None, filename=None):
        super().__init__(path=path, filename=filename)
        self.run()

    def run(self):
        if 'ENERGY_STORAGE' in self.evaluationSets:
            for ID in self.data.keys():
                self.energyStorage(ID)
                print('Vehicle %d: energy storage evaluated' % ID)
            self.fleetEnergyStorage()
            self.calculatedAtts.extend(['energyStorageCritical',
                                        'energyStorageDepleted',
                                        'minEnergyStorageCapacity',
                                        'numVehicles_energyStorageOK',
                                        'numVehicles_energyStorageCritical',
                                        'numVehicles_energyStorageDepleted'])

        if 'TOTAL_ENERGY' in self.evaluationSets:
            for ID in self.data.keys():
                self.energyTotal(ID)
                print('Vehicle %d: energy totals evaluated' % ID)
            self.calculatedAtts.extend(['energyTotalDriving',
                                        'energyTotalPausing',
                                        'energyTotal', 'energyAux',
                                        'energyUnit'])

        if 'CONSUMPTION' in self.evaluationSets:
            # needed to calculate specific consumptions
            if 'TOTAL_ENERGY' in self.evaluationSets:
                for ID in self.data.keys():
                    self.specConsumption(ID)
                    print('Vehicle %d: specific consumptions evaluated' % ID)
                self.calculatedAtts.extend(['specConsumptionOverall',
                                            'specConsumptionDriving'])
            else:
                print('Error: energyTotal needed to calculate specific consumption. Please add \
                              evaluation set \"TOTAL_ENERGY\" to evaluation scheme.')

    def energyStorage(self, ID):
        """
        Evaluate energy storage of vehicle
        """

        # if the lowest SoC is smaller than SoC_min, energyStorage is critical
        self.data[ID]['calculatedAtts_const']['energyStorageCritical'] = \
            self.data[ID]['loggedAtts_time']['energyStorage.SoC'].min() < \
            self.data[ID]['loggedAtts_const']['energyStorage.SoC_min']
        # if the lowest SoC is smaller than SoC_reserve,
        # energyStorage is depleted/impossible
        self.data[ID]['calculatedAtts_const']['energyStorageDepleted'] = \
            self.data[ID]['loggedAtts_time']['energyStorage.SoC'].min() < \
            self.data[ID]['loggedAtts_const']['energyStorage.SoC_reserve']
        # calculate minEnergyStorageCapacity [kWh] using
        # (SoC_high - SoC_low) / (SoC_max - SoC_min)
        SoC_high = self.data[ID]['loggedAtts_time']['energyStorage.SoC'].max()
        SoC_low = self.data[ID]['loggedAtts_time']['energyStorage.SoC'].min()
        SoC_window = self.data[ID]['loggedAtts_const'][
                         'energyStorage.SoC_max'] - \
                     self.data[ID]['loggedAtts_const']['energyStorage.SoC_min']
        self.data[ID]['calculatedAtts_const']['minEnergyStorageCapacity'] = \
            (SoC_high - SoC_low) / SoC_window * \
            self.data[ID]['loggedAtts_const']['energyStorage.capacityNominal']
        # energyStorageDeficit is difference to nominal capacity [kWh]
        self.data[ID]['calculatedAtts_const']['energyStorageDeficit'] = \
            self.data[ID]['loggedAtts_const'][
                'energyStorage.capacityNominal'] - \
            self.data[ID]['calculatedAtts_const']['minEnergyStorageCapacity']

    def fleetEnergyStorage(self):
        """
        Count vehicles which enter each energyStorage state. Currently the count
        is stored in "globalConstants" as there is no class "fleet" yet.
        """

        self.globalConstants['numVehicles_energyStorageOK'] = 0
        self.globalConstants['numVehicles_energyStorageCritical'] = 0
        self.globalConstants['numVehicles_energyStorageDepleted'] = 0
        for n in range(len(self.data)):
            if self.data[n]['calculatedAtts_const']['energyStorageDepleted']:
                self.globalConstants['numVehicles_energyStorageDepleted'] += 1
            elif self.data[n]['calculatedAtts_const']['energyStorageCritical']:
                self.globalConstants['numVehicles_energyStorageCritical'] += 1
            else:
                self.globalConstants['numVehicles_energyStorageOK'] += 1

    def energyTotal(self, ID):
        """
        Evaluate total energy needed for a single vehicle.
        """

        # energyTotalDriving
        self.data[ID]['calculatedAtts_time']['energyTotalDriving'] = \
            self.data[ID]['loggedAtts_time']['energyTraction'] + \
            self.data[ID]['loggedAtts_time']['energyAuxDriving']
        # energyTotalPausing
        self.data[ID]['calculatedAtts_time']['energyTotalPausing'] = \
            self.data[ID]['loggedAtts_time']['energyAuxPausing']
        # energyTotal
        self.data[ID]['calculatedAtts_time']['energyTotal'] = \
            self.data[ID]['calculatedAtts_time']['energyTotalDriving'] + \
            self.data[ID]['calculatedAtts_time']['energyTotalPausing']
        # energyAux
        self.data[ID]['calculatedAtts_time']['energyAux'] = \
            self.data[ID]['loggedAtts_time']['energyAuxPausing'] + \
            self.data[ID]['loggedAtts_time']['energyAuxDriving']
        #        self.data[ID]['energyUnit'] = \
        #   self.data[ID]['energyStorage.energyUnit'][0]  # probably not needed

    def specConsumption(self, ID):
        """
        Evaluate specific vehicle consumptions
        """

        self.data[ID]['calculatedAtts_const']['specConsumptionOverall'] = max(
            self.data[ID]['calculatedAtts_time']['energyTotal']) / max(
            self.data[ID]['loggedAtts_time']['distanceTotal'])
        self.data[ID]['calculatedAtts_const']['specConsumptionDriving'] = max(
            self.data[ID]['calculatedAtts_time']['energyTotalDriving']) / max(
            self.data[ID]['loggedAtts_time']['distanceTotal'])

    def plot_delay(self, IDs=None, show=False, save=True, path='.',
                 baseFilename='Delay_', formats=None):
        """Plot vehicle SoCs with adequate axis limits and with SoC limits as
        dashed lines.

        Temporarily added here until EvaluationVehicle is functional again.
        Can be moved to EvaluationVehicle afterwards.

        DJ
        """
        if not IDs:
            # plot SoC for all vehicles
            IDs = self.data.keys()

        for ID in IDs:
            # Collect data
            currentData = self.data[ID]
            plot_title = currentData['loggedAtts_time']['info'][-1] + \
                '; SoC_end=' + '{:.0f}'.format(currentData['loggedAtts_time']
                ['energyStorage.SoC'][-1]*100) + '%'
            plot_series = currentData['loggedAtts_time']['driver.delay']\
                          /60
            x_min = plot_series.index[0]
            x_max = plot_series.index[-1]

            # Plot dat shite
            plot = LinePlot(title=plot_title, xlabel='Time (hours)',
                            ylabel='Delay (min)', timelabel='hours', show=show)
            plot.addSeries(plot_series)
            if save:
                plot_filename = baseFilename + '{:04d}'.format(ID)
                plot.save(path, plot_filename, formats=formats)

    def plot_SoC(self, IDs=None, show=False, save=True, path='.',
                 baseFilename='SoC_', formats=None):
        """Plot vehicle SoCs with adequate axis limits and with SoC limits as
        dashed lines.

        DJ
        """
        if not IDs:
            # plot SoC for all vehicles
            IDs = self.data.keys()

        for ID in IDs:
            # Collect data
            currentData = self.data[ID]
            plot_title = currentData['loggedAtts_time']['info'][-1] + \
                '; SoC_end=' + '{:.0f}'.format(currentData['loggedAtts_time']
                ['energyStorage.SoC'][-1]*100) + '%'
            plot_series = currentData['loggedAtts_time']['energyStorage.SoC']\
                          * 100
            x_min = plot_series.index[0]
            x_max = plot_series.index[-1]
            SoC_max = currentData['loggedAtts_const']['energyStorage.SoC_max']\
                      * 100
            SoC_min = currentData['loggedAtts_const']['energyStorage.SoC_min']\
                      * 100
            SoC_reserve = currentData['loggedAtts_const']\
                                     ['energyStorage.SoC_reserve'] * 100

            # Plot dat shite
            plot = LinePlot(title=plot_title, xlabel='Time (hours)',
                            ylabel='SoC (%)', timelabel='hours', show=show)
            plot.addSeries(plot_series)
            plot.addSeries([x_min, x_max], [SoC_max, SoC_max], linespec='--k')
            plot.addSeries([x_min, x_max], [SoC_min, SoC_min], linespec='--k')
            plot.addSeries([x_min, x_max], [SoC_reserve, SoC_reserve],
                           linespec='--k')
            # plot.plot(timelabel='hours', show=show)
            plot.axes.set(ylim=[0, 100])
            if save:
                plot_filename = baseFilename + '{:04d}'.format(ID)
                plot.save(path, plot_filename, formats=formats)

    def plot_SoC_with_pauses(self, IDs=None, scheduleList=None, show=False,
                             save=True, path='.', baseFilename='SoC_',
                             formats=None):
        """Plot vehicle SoCs with adequate axis limits and with SoC limits as
        dashed lines.

        DJ
        """
        if not IDs:
            # plot SoC for all vehicles
            IDs = self.data.keys()

        for ID in IDs:
            # Collect SoC data from eval
            currentData = self.data[ID]
            plot_title = currentData['loggedAtts_time']['info'][-1] + \
                '; SoC_end=' + '{:.0f}'.format(currentData['loggedAtts_time']
                ['energyStorage.SoC'][-1]*100) + '%'
            plot_series = currentData['loggedAtts_time']['energyStorage.SoC']\
                          * 100
            extractHours = lambda pandasObj: (pandasObj.day-1)*24\
                + pandasObj.hour + pandasObj.minute/60\
                + pandasObj.second/3600

            #ind = plot_series.index
            hours = pd.Series(extractHours(plot_series.index))
            x_min = hours[hours.index[0]]
            x_max = hours[hours.index[-1]]
            # x_min = plot_series.index[0]
            # x_max = plot_series.index[-1]
            SoC_max = currentData['loggedAtts_const']['energyStorage.SoC_max']\
                      * 100
            SoC_min = currentData['loggedAtts_const']['energyStorage.SoC_min']\
                      * 100
            SoC_reserve = currentData['loggedAtts_const']\
                                     ['energyStorage.SoC_reserve'] * 100

            # Plot dat shite
            # Instantiate plot
            plot = LinePlot(title=plot_title, xlabel='Time (hours)',
                            ylabel='SoC (%)', show=show)
            # plot = LinePlot(title=plot_title, xlabel='Time (hours)',
            #                 ylabel='SoC (%)', timelabel='hours', show=show)
            plot.addSeries(hours, plot_series.values)
            plot.addSeries([x_min, x_max], [SoC_max, SoC_max], linespec='--k')
            plot.addSeries([x_min, x_max], [SoC_min, SoC_min], linespec='--k')
            plot.addSeries([x_min, x_max], [SoC_reserve, SoC_reserve],
                           linespec='--k')

            # Collect schedule (pause) data

            # This is an ugly hack: scheduleID is constant, but logged at every
            # timestep because of bug in DataLogger. Thus, we take the first
            # value:
            scheduleID = currentData['loggedAtts_time']['driver.schedule.ID'][0]
            if scheduleList is not None:
                for schedule in scheduleList:
                    if schedule.ID == scheduleID:
                        thisSchedule = schedule
                        break  # no need to continue iterating
            try:
                pauses = []
                # arrivals = []
                # pauseDurations = []
                for trip in thisSchedule.tripList:
                    # only evaluate end of trip:
                    leg = trip.legList[-1]
                    # for leg in trip.legList:
                    if leg.pause > 0:
                        beginPause = pd.to_datetime(
                            abs(leg.departureTime) + leg.duration,
                            unit='s',
                            origin=pd.Timestamp('2018-01-01'))
                        endPause = pd.to_datetime(
                            abs(leg.departureTime) + leg.duration + leg.pause,
                            unit='s',
                            origin=pd.Timestamp('2018-01-01'))
                        beginPause = extractHours(beginPause)
                        endPause = extractHours(endPause)
                        pauseLength = endPause - beginPause
                        pauses.append((beginPause, pauseLength))
                        # pauses.append(beginPause)
                        # pauses.append(endPause)
                    # arrivals.append(abs(leg.departureTime) + leg.duration)
                    # pauseDurations.append(leg.pause)
                plot.axes.broken_barh(pauses, (0, 100), color=(1, 0, 0, 0.2), edgecolor='None')
                # for x in pauses:
                #     plot.axes.axvline(x, ymin=0, ymax=100, linestyle='dashed')
            except NameError:
                # thisSchedule doesn't exist. Do nothing
                pass


            plot.axes.set(ylim=[0, 100])
            def convertHours(x, pos):
                if x >= 24:
                    tick = x - 24
                else:
                    tick = x
                return '{:02.0f}'.format(tick)

            plot.axes.xaxis.set_major_formatter(matplotlib.ticker.
                                                FuncFormatter(convertHours))
            plot.axes.xaxis.set_major_locator(matplotlib.ticker.
                                              MultipleLocator(base=2))
            plot.axes.xaxis.set_minor_locator(matplotlib.ticker.
                                              MultipleLocator(base=1))
            # plot.axes.grid(False)
            if save:
                plot_filename = baseFilename + '{:04d}'.format(ID)
                plot.save(path, plot_filename, formats=formats)

    def plot_energyBalancePie(self, show=False, save=True, path='.',
                              filename='EnergyBalance', formats=None):
        """Plot energy balance of vehicles/schedules as a pie chart.

        Dominic Jefferies"""
        colors = ['mediumaquamarine', 'khaki', 'darksalmon']
        sizes = [self.globalConstants['numVehicles_energyStorageOK'],
                 self.globalConstants['numVehicles_energyStorageCritical'],
                 self.globalConstants['general']['numVehicles_energyStorageDepleted']]

        labels = ['OK', 'critical', 'impossible']
        title = 'Energy Balance'

        plot = PiePlot(sizes, labels=labels, title=title, colors=colors,
                       show=show)

        if save:
            plot.save(path, filename, formats=formats)


class EvaluationSimpleDepot(Evaluation):
    def __init__(self, path=None, filename=None):
        super().__init__(path=path, filename=filename)

    def plot_numVehicles(self, IDs=None, show=False, save=True, path='.',
                 baseFilename='numVehicles_Depot_', formats=None):
        """Plot number of vehicles in depot/in service over time. Can be
        applied to a DepotStack or to a list of Depots.

        DJ
        """
        if not IDs:
            # plot data for all depots
            IDs = self.data.keys()

        for ID in IDs:
            # Collect data
            currentData = self.data[ID]
            plot_title = None

            plot_series_names = ['numVehiclesInService',
                                 'numVehiclesCharging', 'numVehiclesReady']

            plot_series_labels = ['In service', 'Charging',
                                  'Charging finished']

            plot_series = []

            for name in plot_series_names:
                plot_series.append(currentData['loggedAtts_time'][name])
            # x_min = plot_series.index[0]
            # x_max = plot_series.index[-1]

            # Plot dat shite
            plot = LinePlot(title=plot_title, xlabel='Time (hours)',
                            ylabel='Number of vehicles', timelabel='hours', show=show)
            for i, series in enumerate(plot_series):
                plot.addSeries(series.index.values, series, label=plot_series_labels[i], step=True)

            plot.axes.legend(loc='upper left', frameon=True)
            plot.axes.yaxis.set_major_locator(matplotlib.ticker.
                                              MultipleLocator(base=2))

            if save:
                plot_filename = baseFilename + '{:04d}'.format(ID)
                plot.save(path, plot_filename, formats=formats)

class EvaluationDepot(Evaluation):
    pass


class EvaluationChargingFacility(Evaluation):
    pass


class EvaluationBus(EvaluationVehicle):
    pass


class EvaluationTrain(EvaluationVehicle):
    pass


class Plot:
    """Base class for plots.

    Dominic Jefferies"""

    def __init__(self, title=None, figureSize=None, show=False, lineWidth=1.0):
        self.title = title
        self.figureSize = figureSize
        self.figure, self.axes = plt.subplots(figsize=self.figureSize
            if self.figureSize else globalConstants['general'][
                'DEFAULT_PLOT_SIZE'], linewidth=lineWidth)
        if show:
            plt.interactive(True)
            plt.show()
        else:
            plt.interactive(False)  # just to be sure...

    def save(self, path, filename, formats=None):
        if not formats:
            # default, save as png
            self.figure.savefig(path + '\\' + filename + '.png')
        else:
            if 'png' in formats:
                self.figure.savefig(path + '\\' + filename + '.png')
            if 'pdf' in formats:
                self.figure.savefig(path + '\\' + filename + '.pdf')
        plt.close(self.figure)


class LinePlot(Plot):
    """Class for a line plot that allows easy plotting of time series. After
    creating a LinePlot object, any modification can be made by manipulating the
    self.figure and self.axes attributes as per the matplotlib documentation
    (e.g., set axis limits, ...).

    Dominic Jefferies
    """

    def __init__(self, title=None, xlabel=None, ylabel=None, timelabel=None,
                 figureSize=None, show=False, lineWidth=1.0):
        super().__init__(title=title, figureSize=figureSize, show=show, lineWidth=lineWidth)
        self.linewidth = lineWidth
        self.xlabel = xlabel
        self.ylabel = ylabel
        if title is not None:
            self.axes.set(xlabel=self.xlabel, ylabel=self.ylabel, title=self.title)
        else:
            self.axes.set(xlabel=self.xlabel, ylabel=self.ylabel)
        self.axes.grid(b=True, zorder=0, which='both')
        self.axes.legend(frameon=True)

        if timelabel == 'hours':
            majorLocator = mdates.HourLocator(byhour=range(0, 24, 2))
            minorLocator = mdates.HourLocator()
            majorFormat = mdates.DateFormatter('%H')
            self.axes.xaxis.set_major_locator(majorLocator)
            self.axes.xaxis.set_minor_locator(minorLocator)
            self.axes.xaxis.set_major_formatter(majorFormat)

        # other timelabel configurations to be added when needed

    def addSeries(self, *data, xerror = None, yerror = None, linespec=None,
                  label=None, step=False):
        """Add a series to the plot. Data can be specified as an x, y pair
        or as y values only. If you supply a Pandas series as a single
        argument, the index will be used as x values. When step=True,
        you MUST supply x and y arguments (matplotlib limitation). To plot a
        Pandas timeseries ts as a step plot, pass it as
        'ts.index.values, ts'.

        If a label is specified, it will automatically be added as a legend
        entry. Series with label=None are omitted in the legend.
        """
        if len(data) < 1:
            raise ArgumentError('You must provide at least one argument '
                                'containing plot data.')

        if step and len(data) < 2:
            raise ArgumentError('When using step, always supply x and y data.')

        xdata = data[0] if len(data) > 1 else None
        ydata = data[0] if len(data) == 1 else data[1]

        if xerror or yerror:
            if step:
                raise ArgumentError('Step plots with error bars currently '
                                    'not possible.')
            else:
                if xdata is not None:
                    if linespec:
                        self.axes.errorbar(xdata, ydata, linespec, label=label)
                    else:
                        self.axes.errorbar(xdata, ydata, label=label)
                else:
                    raise ArgumentError('When plotting error bars, x and y '
                                        'data must each be supplied '
                                        'explicitly.')
        else:
            if step:
                if linespec:
                    self.axes.step(xdata, ydata, linespec, label=label,
                                   where='post', linewidth = self.linewidth)
                else:
                    self.axes.step(xdata, ydata, label=label, where='post', linewidth = self.linewidth)
            else:
                if xdata is not None:
                    if linespec:
                        self.axes.plot(xdata, ydata, linespec, label=label, linewidth=self.linewidth)
                    else:
                        self.axes.plot(xdata, ydata, label=label, linewidth=self.linewidth)
                else:
                    if linespec:
                        self.axes.plot(ydata, linespec, label=label, linewidth=self.linewidth)
                    else:
                        self.axes.plot(ydata, label=label, linewidth=self.linewidth)

    def addRectangles(self, xdata, ydata, height=0.5, color='red', fill=True):
        """ adds rectangles according to values specified in xdata, ydata

        Tobias Altmann"""
        def addRectangle(point, width, height, color, fill):
            self.axes.add_patch(
                patches.Rectangle(
                    point, width, height, fill=fill, color=color
                )
            )

        self.addSeries(xdata, ydata, step=True)
        values = [y for y in ydata if not math.isnan(y)]
        maxVal = max(values) if values else None
        if maxVal:
            self.axes.set_ylim([0,maxVal+1])
        nan_indices = sorted([i for i in range(len(ydata)) if math.isnan(ydata[i])])
        value_indices = sorted(list(set(range(len(ydata))) - set(nan_indices)))
        while value_indices:
            index = value_indices.pop(0)
            nan_indices = [i for i in nan_indices if i > index]
            yval = ydata[index]
            xval = xdata[index]
            index = nan_indices.pop(0) if nan_indices else len(xdata) - 1
            width = xdata[index] - xval
            addRectangle((xval, yval - height / 2), width, height, color, fill)
            value_indices = [i for i in value_indices if i > index]
        self.axes.lines.pop(0)

class PiePlot(Plot):
    """Class for a pie plot with automatic pie segment labelling of the form
    'x (y%)'.

    Dominic Jefferies
    """

    def __init__(self, pieSizes, labels=None, colors=None, title=None,
                 figureSize=None, show=False):
        super().__init__(title=title, figureSize=figureSize, show=show)
        self.pieSizes = pieSizes
        self.labels = labels
        self.axes.set(title=self.title)
        if colors is not None:
            self.axes.pie(self.pieSizes, colors=colors,
                          autopct=self.pieLabels(self.pieSizes),
                          startangle=90, counterclock=False)
        else:
            self.axes.pie(self.pieSizes, autopct=self.pieLabels(self.pieSizes),
                          startangle=90, counterclock=False)
        self.axes.axis('equal')  # make pie a circle, not an oval
        if labels:
            self.axes.legend(labels, frameon=True)

    @staticmethod
    def pieLabels(values):
        def autopct(pct):
            total = sum(values)
            val = int(round(pct * total / 100.0))
            return '{v:d} ({p:.0f}%)'.format(p=pct, v=val)
        return autopct


class SimulationReport:
    def __init__(self, title=None, eval_vehicles=None, eval_infra=None,
                 vehicleParams_dict=None, vehicleStack=None,
                 tripReport_post_process=None, chargingInterfaces=None,
                 globalConstants=None, dictionary=None):
        self.title = title
        self.eval_vehicles = eval_vehicles
        self.eval_infra = eval_infra
        self.chargingInterfaces = chargingInterfaces
        self.vehicleParams_dict = vehicleParams_dict
        self.globalConstants = globalConstants
        self.globalReportData = dict()
        self.vehicleReportData = dict()
        self.dictionary = dictionary

        if vehicleStack is not None:
            # Generate trip reports
            for vehicle in vehicleStack:
                vehicle_df = pd.DataFrame.from_dict(vehicle.driver.tripReport,
                                                    orient='index')
                if tripReport_post_process is not None:
                    vehicle_df = tripReport_post_process(vehicle_df)
                self.vehicleReportData.update(
                    {vehicle.driver.schedule.ID: vehicle_df})

        if vehicleParams_dict is not None:
            vehicleData = dict()
            for name, params in vehicleParams_dict.items():
                vParams = {translate('Name', dictionary):
                               params['name'],
                           translate('Specific drivetrain consumption', dictionary):
                               [params['specificDriveConsumption'], 'kWh/km'],
                           translate('Battery capacity', dictionary):
                               [params['energyStorageParams']['capacityNominal'], 'kWh'],
                           translate('SoC max', dictionary):
                               [params['energyStorageParams']['SoC_max']],
                           translate('SoC min', dictionary):
                               [params['energyStorageParams']['SoC_min']],
                           translate('SoC reserve', dictionary):
                               [params['energyStorageParams']['SoC_reserve']],
                           translate('SoH', dictionary):
                               [params['energyStorageParams']['SoC_reserve']],
                           translate('Max. C rate', dictionary):
                               [params['energyStorageParams']['C_rate']],
                           translate('Max. charging power', dictionary):
                               [params['energyStorageParams']['capacityNominal']*
                                    params['energyStorageParams']['C_rate'], 'kW'],
                           translate('HVAC system', dictionary):
                               [params['airConditioning']]
                           }
                vehicleData.update({translate(name, dictionary): vParams})
            self.globalReportData.update(
                {translate('Vehicles', dictionary): vehicleData})
        else:
            vehicleData = None

        if chargingInterfaces is not None:
            chargingInterfaceData = dict()
            for interface in chargingInterfaces:
                name = interface.name
                cParams = {
                    translate('Dead time before', dictionary):
                        [interface.deadTimeBefore, 's'],
                    translate('Dead time after', dictionary):
                        [interface.deadTimeAfter, 's'],
                    translate('Max. power', dictionary):
                        [interface.maxPower, 'kW'],
                    translate('Efficiency', dictionary):
                        [interface.efficiency]
                }
                chargingInterfaceData.update({translate(name, dictionary): cParams})
            self.globalReportData.update(
                {translate('Charging interfaces', dictionary): chargingInterfaceData})

        if eval_infra is not None:
            infraData = dict()
            for cf in eval_infra.data.values():
                name = cf['loggedAtts_const']['location.name']
                cfParams = {
                    translate('Charging interface', dictionary):
                        [cf['loggedAtts_const']['interface.name']],
                    translate('Max. number of vehicleStack', dictionary):
                        [max(cf['loggedAtts_time']['numVehicles'])]
                }
                infraData.update(
                    {translate(name, dictionary): cfParams})
            self.globalReportData.update(
                {translate('Charging infrastructure',
                           dictionary): infraData})

        if globalConstants is not None:
            globalData = {
                translate('Ambient temperature', dictionary):
                    [globalConstants['general']['AMBIENT_TEMPERATURE'], '°C'],
                translate('Always wait for charging', dictionary):
                    [globalConstants['network']['WAIT_FOR_CHARGING']]
            }
            self.globalReportData.update({translate('Other parameters', dictionary):
                                    globalData})

    def save_xls(self, path, file):
        # Save global report data
        xlspath = path + '\\' + file
        workbook = xlsxwriter.Workbook(xlspath)
        worksheet = workbook.add_worksheet(translate(self.title, self.dictionary))

        heading_format = workbook.add_format({'bold': True, 'font_size': 16})
        subheading_format = workbook.add_format({'bold': True, 'font_size': 12})
        bold_format = workbook.add_format({'bold': True})

        row = 0
        col = 0

        worksheet.write(row, col, self.title, heading_format)
        row += 2

        for cat, cat_data in self.globalReportData.items():
            col = 0
            worksheet.write(row, col, cat, subheading_format)
            row += 2
            for subcat, params in cat_data.items():
                col = 0
                if type(params) == dict:
                    # we have a further layer to unpack
                    worksheet.write(row, col, subcat, bold_format)
                    row += 1
                    for name, values in params.items():
                        col = 0
                        worksheet.write(row, col, name)
                        col += 1
                        if type(values) == str:
                            worksheet.write(row, col, values)
                        else:
                            for value in values:
                                worksheet.write(row, col, value)
                                col += 1
                        row += 1
                    row += 1
                elif type(params) == list:
                    # only one layer
                    worksheet.write(row, col, subcat)
                    col += 1
                    for value in params:
                        worksheet.write(row, col, value)
                        col += 1
                    row += 1
            row += 1
        workbook.close()

        # Save vehicle trip reports
        workbook = load_workbook(xlspath)
        writer = pd.ExcelWriter(xlspath, engine='openpyxl')
        writer.book = workbook
        writer.sheets = dict((sheet.title, sheet) for sheet in workbook.worksheets)
            #[sheet.get_name() for sheet in workbook.worksheets()]
        for vehicleID in self.vehicleReportData.keys():
            self.vehicleReportData[vehicleID].to_excel(writer,
                sheet_name=str(vehicleID))
        writer.save()